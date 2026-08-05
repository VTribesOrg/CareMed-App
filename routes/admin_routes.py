from flask import Blueprint, render_template, url_for, redirect, flash, request, jsonify, current_app, send_from_directory, abort
from flask_login import current_user
from extensions import db, limiter, csrf
from sqlalchemy.orm import joinedload
from sqlalchemy import func, or_, and_
from flask_login import login_required
from functools import wraps
from models.product import Product, InventoryLog, Transaction, Purchase, Payment, Rental, RefillTransaction, RentalInvoice, Expense, PaymentProof, TankStatus
from models.customer import Customer
from models.users import User, SecurityLog, BlockedIP, Permission
from flask_mail import Message
from flask import current_app
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from dateutil.relativedelta import relativedelta
from sqlalchemy.orm import joinedload, selectinload
import os   
import uuid
import random, string
from utils.backup import create_backup, get_all_backups
from flask import send_file, Response, stream_with_context
import json
import time


admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

def admin_or_staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):

        authorized_roles = ['Administrator', 'Staff']

        if not current_user.is_authenticated or current_user.role.strip() not in authorized_roles:
            try:
                log = SecurityLog(
                    ip_address=request.remote_addr,
                    event_type="Unauthorized Access",
                    description=f"User with role '{getattr(current_user, 'role', 'None')}' tried to access: {request.path}",
                    user_id=current_user.id if current_user.is_authenticated else None,
                    user_email=current_user.email if current_user.is_authenticated else None,
                    user_agent=request.headers.get('User-Agent', 'Unknown')[:255],
                    severity='High',
                    is_suspicious=True
                )
                db.session.add(log)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Failed to log unauthorized access: {e}")
                
            flash("Unauthorized access.", "error")
            return redirect(url_for('user.homepage'))
            
        return f(*args, **kwargs)
    return decorated_function

def permission_required(permission_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('auth.login'))
            
            if getattr(current_user, 'role', '') == 'Administrator':
                return f(*args, **kwargs)
            
            user_perms = Permission.query.filter_by(user_id=current_user.id).first()
            

            if user_perms and getattr(user_perms, permission_name, False):
                return f(*args, **kwargs)

            flash("You do not have permission to access this feature.", "danger")
            return redirect(request.referrer or url_for('admin.dashboard'))
            
        return decorated_function
    return decorator

from sqlalchemy import func

def get_cogs_category(product):
    """Smart category assignment for auto-generated Cost of Sales expenses."""
    if product.is_refillable:
        return 'Oxygen Refill Cost'
    et = (product.equipment_type or '').lower()
    name = (product.name or '').lower()
    if 'medical supplies' in et or 'medical supplies' in name:
        return 'Medical Supplies Purchase'
    if 'supplies' in et or 'supplies' in name:
        return 'Medical Supplies Purchase'
    return 'Equipment Purchase'

COST_OF_SALES_CATEGORIES = {
    'Oxygen Refill Cost',
    'Equipment Purchase',
    'Medical Supplies Purchase',
    'Inventory Purchase',
}
 
OPERATING_EXPENSE_CATEGORIES = {
    'Utilities',
    'Rent',
    'Salaries',
    'Maintenance',
    'Government Fees',
    'Insurance',
    'Transportation',
    'Office Supplies',
    'Other',
}

ALL_EXPENSE_CATEGORIES_ORDERED = [
    # Cost of Sales
    'Oxygen Refill Cost',
    'Equipment Purchase',
    'Medical Supplies Purchase',
    'Inventory Purchase',
    # Operating Expenses
    'Utilities',
    'Rent',
    'Salaries',
    'Maintenance',
    'Government Fees',
    'Insurance',
    'Transportation',
    'Office Supplies',
    'Other',
]

@admin_bp.route('/dashboard')
@login_required
@admin_or_staff_required
def dashboard():
    total_sales = db.session.query(
        func.sum(Transaction.amount_paid)
    ).filter_by(transaction_type='Sale').scalar() or 0

    total_rentals = db.session.query(
        func.sum(Payment.amount)
    ).join(Transaction).filter(
        Transaction.transaction_type == 'Rental',
        Payment.status == 'Completed'
    ).scalar() or 0

    # Calculate total refill profit
    total_refill_profit = db.session.query(
        func.sum(RefillTransaction.total_revenue)
    ).scalar() or 0

    active_rentals_count = Rental.query.filter_by(status='Active').count()
    
    total_expenses = db.session.query(
        func.sum(Expense.amount)
    ).scalar() or 0

    product_inventory = db.session.query(
        func.sum(Product.stock)
    ).filter(
        Product.is_active == True,
        Product.is_refillable == False
    ).scalar() or 0

    # ── Aggregate Tank Statuses ──────────────────────────────────────────
    raw_tank_data = TankStatus.query.join(Product).filter(Product.is_active == True).all()
    
    tank_aggregation = {}
    total_tank_count = 0
    
    for tank in raw_tank_data:
        # Create a unique key based on name and size
        key = (tank.product.name, tank.product.size)
        
        if key not in tank_aggregation:
            tank_aggregation[key] = {
                'name': tank.product.name,
                'size': tank.product.size,
                'product_id': tank.product.id,
                'total_owned': 0,
                'rented_out': 0,
                'full_in_stock': 0,
                'empty_in_stock': 0
            }
        
        tank_aggregation[key]['total_owned'] += tank.total_owned
        tank_aggregation[key]['rented_out'] += tank.rented_out
        tank_aggregation[key]['full_in_stock'] += tank.full_in_stock
        tank_aggregation[key]['empty_in_stock'] += tank.empty_in_stock
        
        # Add to total for inventory calculation
        total_tank_count += (tank.full_in_stock + tank.empty_in_stock + tank.rented_out)

    combined_tank_statuses = list(tank_aggregation.values())
    total_inventory = product_inventory + total_tank_count

    low_stock_count = Product.query.filter(
        Product.is_active == True,
        Product.is_refillable == False,
        Product.stock <= 5
    ).count()

    all_products = Product.query.filter(
        Product.is_refillable == False, 
        Product.is_active == True,
        Product.condition.in_(['Brand New', 'Used'])
    ).all()
    
    active_rented_units = db.session.query(
        Rental.product_id,
        func.sum(Rental.quantity - Rental.quantity_returned)
    ).filter(Rental.status == 'Active').group_by(Rental.product_id).all()
    
    rented_map = {prod_id: count for prod_id, count in active_rented_units}

    assets_aggregation = {}
    for prod in all_products:
        prod_name = prod.name
        if prod_name not in assets_aggregation:
            assets_aggregation[prod_name] = {
                'name': prod_name,
                'total_stock': 0,
                'rented_count': 0,
                'used_count': 0,
                'brand_new_count': 0
            }
            
        assets_aggregation[prod_name]['total_stock'] += (prod.stock or 0)
        assets_aggregation[prod_name]['rented_count'] += rented_map.get(prod.id, 0)

        condition_str = (prod.condition or "").strip()
        if condition_str == "Brand New":
            assets_aggregation[prod_name]['brand_new_count'] += (prod.stock or 0)
        elif condition_str == "Used":
            assets_aggregation[prod_name]['used_count'] += (prod.stock or 0)

    standard_assets = list(assets_aggregation.values())

    recent_logs = InventoryLog.query.order_by(
        InventoryLog.created_at.desc()
    ).limit(5).all()

    security_alerts = SecurityLog.query.order_by(
        SecurityLog.created_at.desc()
    ).limit(3).all()
    
    active_oxygen_rentals = Rental.query.join(Product).filter(
        Rental.status == 'Active',
        Product.is_refillable == True
    ).all()

    return render_template(
        'admin/dashboard.html',
        total_sales=total_sales,
        total_rentals=total_rentals,
        total_refill_profit=total_refill_profit, # New variable added
        active_rentals_count=active_rentals_count,
        total_inventory=total_inventory,
        low_stock_count=low_stock_count,
        tank_statuses=combined_tank_statuses,
        standard_assets=standard_assets,
        recent_logs=recent_logs,
        security_alerts=security_alerts,
        total_expenses=total_expenses,
        active_oxygen_rentals=active_oxygen_rentals
    )
    
@admin_bp.route('/request-refill', methods=['POST'])
@login_required
@admin_or_staff_required
def request_refill():
    product_id = request.form.get('product_id')
    quantity = request.form.get('quantity', type=int)
    rental_id = request.form.get('rental_id')

    if not product_id or not quantity or quantity <= 0:
        flash("It looks like some information is missing. Please check your request and try again.", "error")
        return redirect(url_for('admin.dashboard'))

    try:
        tank = TankStatus.query.filter_by(product_id=product_id).first()
        if not tank:
            flash("We couldn't find this specific tank in our records. Please contact support.", "error")
            return redirect(url_for('admin.dashboard'))

        if tank.full_in_stock < quantity:
            flash(f"Oops! We only have {tank.full_in_stock} full tanks ready. Please adjust your request to a lower amount.", "error")
            return redirect(url_for('admin.dashboard'))

        ref_no = "Manual"
        if rental_id:
            rental = Rental.query.get(rental_id)
            if rental and rental.transaction:
                ref_no = rental.transaction.reference_no

        tank.full_in_stock -= quantity
        tank.empty_in_stock += quantity

        log = InventoryLog(
            product_id=product_id,
            action="Refill Request Created",
            quantity=quantity,
            note=f"Moved {quantity} units from Full to Empty. Rental Ref: {ref_no}",
            user_id=current_user.id,
            user_name=current_user.full_name
        )
        
        db.session.add(log)
        db.session.commit()
        
        flash(f"Success! {quantity} tank(s) have been marked as empty and are now ready for refill.", "success")

    except Exception as e:
        db.session.rollback()
        flash("We ran into a slight issue updating the inventory. Please try again or contact the technical team if this persists.", "error")

    return redirect(url_for('admin.dashboard'))
   
@admin_bp.route('/process-refill-transaction', methods=['POST'])
@login_required
@admin_or_staff_required
def process_refill_transaction():
    buyer_type = request.form.get('refill_buyer_type')
    tank_size = request.form.get('tank_size') # Fallback or select value depending on implementation
    quantity = request.form.get('quantity', type=int)
    serial_input = request.form.get('serial_numbers', '')
    serial_list = [sn.strip() for sn in serial_input.split(',') if sn.strip()]
    amount = request.form.get('amount', type=Decimal)
    
    REFILL_COST = Decimal("50.00") 

    if not amount or not quantity or quantity <= 0:
        flash("Transaction failed: Please ensure all fields are filled correctly.", "error")
        return redirect(url_for('admin.dashboard'))

    if buyer_type == 'registered' and len(serial_list) != quantity:
        flash(f"Data Mismatch: You specified {quantity} tank(s), but only entered {len(serial_list)} serial number(s).", "error")
        return redirect(url_for('admin.dashboard'))

    try:
        customer_id = None
        display_name = "Walk-in Customer"
        product = None
        
        if buyer_type == 'registered':
            customer_id = request.form.get('refill_customer_id')
            customer = Customer.query.get(customer_id)
            if not customer:
                flash("System Error: Could not verify the registered customer.", "error")
                return redirect(url_for('admin.dashboard'))
            display_name = customer.full_name
            
            # Retrieve selected active rental serial being swapped out (if any)
            swapped_serial = request.form.get('swapped_rental_serial')
            
            if swapped_serial:
                # Find active rental associated with this serial number to swap
                active_rental = Rental.query.filter_by(serial_number=swapped_serial, status="Active").first()
                if active_rental:
                    product = active_rental.product
                    # Update rental record serial number swap
                    active_rental.serial_number = serial_list[0] if serial_list else active_rental.serial_number
            
            if not product:
                # Fallback to finding product via form selection if not derived from active rental
                selected_tank_val = request.form.get('tank_size_select', '')
                product_name = selected_tank_val.split(' - ')[0] if ' - ' in selected_tank_val else selected_tank_val
                product = Product.query.filter_by(name=product_name, is_refillable=True).first()

            if product:
                tank_status = TankStatus.query.filter_by(product_id=product.id).first()
                if not tank_status or tank_status.full_in_stock < quantity:
                    flash(f"Inventory Alert: Insufficient full tanks in stock for {product.name}.", "error")
                    return redirect(url_for('admin.dashboard'))
                
                # Minus full, add empty
                tank_status.full_in_stock -= quantity
                tank_status.empty_in_stock += quantity
        else:
            display_name = request.form.get('unregistered_customer_name', '').strip() or "Walk-in Customer"
            selected_tank_val = request.form.get('tank_size_select', '')
            product_name = selected_tank_val.split(' - ')[0] if ' - ' in selected_tank_val else selected_tank_val
            product = Product.query.filter_by(name=product_name, is_refillable=True).first()

        new_transaction = RefillTransaction(
            product_id=product.id if product else None,
            customer_id=customer_id,
            walk_in_name=display_name if buyer_type != 'registered' else None,
            walk_in_tank_size=tank_size if not product else None,
            quantity=quantity,
            total_revenue=amount,
            refill_cost_per_unit=REFILL_COST,
            serial_numbers=", ".join(serial_list) if buyer_type == 'registered' else "N/A",
            processed_by_id=current_user.id
        )
        db.session.add(new_transaction)

        log = InventoryLog(
            product_id=product.id if product else None,
            action="Refill Service Completed & Swapped",
            quantity=quantity,
            note=f"Processed refill swap for {display_name}. Total: {amount} PHP.",
            user_id=current_user.id,
            user_name=current_user.full_name
        )
        db.session.add(log)
        
        db.session.commit()
        flash(f"Refill and tank swap successfully processed for {display_name}.", "success")

    except Exception as e:
        db.session.rollback()
        flash("An unexpected error occurred while saving the transaction. Please try again.", "error")

    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/customers')
@login_required
@admin_or_staff_required
@permission_required('can_manage_customers')
def customers():
    page         = request.args.get('page', 1, type=int)
    limit        = request.args.get('limit', 10, type=int)
    search_query = request.args.get('q', '').strip()
    filter_type  = request.args.get('filter', '')         
 
    query = Customer.query
 
    if search_query:
        query = query.filter(or_(
            Customer.first_name.ilike(f"%{search_query}%"),
            Customer.last_name.ilike(f"%{search_query}%"),
            Customer.contact_number.ilike(f"%{search_query}%")
        ))
 
    if filter_type == 'overdue_invoices':
        from models.product import RentalInvoice, Rental as RentalModel
        today = date.today()
        overdue_customer_ids = db.session.query(
            RentalModel.customer_id
        ).join(
            RentalInvoice, RentalInvoice.rental_id == RentalModel.id
        ).filter(
            RentalInvoice.status != 'Paid',
            RentalInvoice.service_period_start < today
        ).distinct().subquery()
 
        query = query.filter(Customer.id.in_(overdue_customer_ids))
 
    elif filter_type == 'pending_id':

        query = query.filter(
            Customer.is_id_verified == False,
            Customer.valid_id_path != None
        )
 
    elif filter_type == 'inactive_open':
        open_txn_customer_ids = db.session.query(
            Transaction.customer_id
        ).filter(
            Transaction.status == 'Open'
        ).distinct().subquery()
 
        query = query.filter(
            Customer.is_active == False,
            Customer.id.in_(open_txn_customer_ids)
        )
    
 
    pagination = query.order_by(Customer.last_name.asc()).paginate(
        page=page,
        per_page=limit,
        error_out=False
    )
 
    return render_template(
        "admin/customers.html",
        customers=pagination.items,
        pagination=pagination,
        search_query=search_query,
        current_limit=limit,
        current_filter=filter_type      
    )

@admin_bp.route('/get_customer/<int:id>')
@login_required
@admin_or_staff_required
def get_customer(id):
    try:
        customer = Customer.query.options(
            db.joinedload(Customer.user)
        ).get(id)

        if not customer:
            return jsonify({
                "status": "error",
                "message": "Customer record not found."
            }), 404

        profile_url = None
        if customer.user and customer.user.profile_path:
            if customer.user.profile_path.startswith(('http://', 'https://')):
                profile_url = customer.user.profile_path
            else:
                profile_url = url_for('static', filename=customer.user.profile_path)

        return jsonify({
            "status": "success",
            "data": {
                "id": customer.id,
                "first_name": customer.first_name or "",
                "last_name": customer.last_name or "",
                "full_name": customer.full_name or "", 
                
                "contact_number": customer.contact_number or "",
                "secondary_contact_number": customer.secondary_contact_number or "",
                "home_address": customer.home_address or "",
                
                "gender": customer.gender or "",
                
                "is_active": customer.is_active,
                "is_id_verified": customer.is_id_verified,
                
                "primary_id_type": customer.primary_id_type or "",
                "secondary_id_type": customer.secondary_id_type or "",
                
                "valid_id_path": url_for('static', filename=customer.valid_id_path) if customer.valid_id_path else None,
                "secondary_id_path": url_for('static', filename=customer.secondary_id_path) if customer.secondary_id_path else None,
                "proof_of_billing_path": url_for('static', filename=customer.proof_of_billing_path) if customer.proof_of_billing_path else None,
                
                "id_uploaded_at": customer.id_uploaded_at.strftime('%b %d, %Y %I:%M %p') if customer.id_uploaded_at else "Never",
                
                "has_online_account": True if customer.user_id else False,
                "email": customer.user.email if customer.user else "Walk-in / No Email",
                "profile_path": profile_url,
                
                "created_at": customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else None
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error fetching customer {id}: {str(e)}")
        return jsonify({
            "status": "error",
            "message": "An internal server error occurred while retrieving customer data."
        }), 500
        
        
@admin_bp.route('/customers/<int:id>')
@login_required
@admin_or_staff_required
def customer_details(id):
    customer = Customer.query.get_or_404(id)

    transactions = (Transaction.query .filter_by(customer_id=id) .order_by(Transaction.created_at.desc()).all())

    return render_template('admin/customer_details.html', customer=customer, transactions=transactions)

@admin_bp.route('/customers/<int:id>/verify', methods=['POST'])
@login_required
@admin_or_staff_required
def verify_customer(id):
    customer = Customer.query.get_or_404(id)
    
    if customer.is_id_verified:
        flash(f"Notice: {customer.full_name} is already a verified user.", "info")
        return redirect(url_for('admin.customer_details', id=id))

    try:
        customer.is_id_verified = True
        db.session.commit()
        flash(f"Identity Verification Complete: {customer.full_name} has been successfully verified.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Verification Failed for ID {id}: {str(e)}") 
        flash("An internal error occurred while processing the verification. Please refresh and try again.", "danger")
        
    return redirect(url_for('admin.customer_details', id=id))

@admin_bp.route('/customers/<int:id>/unverify', methods=['POST'])
@login_required
@admin_or_staff_required
def unverify_customer(id):
    customer = Customer.query.get_or_404(id)
    
    try:
        customer.is_id_verified = False
        db.session.commit()
        flash(f"Verification Revoked: {customer.full_name}'s identity status has been reset to unverified.", "warning")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Revoke Error for ID {id}: {str(e)}")
        flash("Action Failed: System encountered an error while revoking verification. Please try again.", "danger")
        
    return redirect(url_for('admin.customer_details', id=id))


@admin_bp.route('/admin/add-customer', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_manage_customers')
def add_customer():

    if current_user.role not in ['Administrator', 'Staff']:
        flash("Unauthorized access.", "error")
        return redirect(url_for('user.homepage'))

    first_name = request.form.get('first_name', '').strip().title()
    last_name = request.form.get('last_name', '').strip().title()
    contact_number = request.form.get('contact_number', '').strip()
    
    secondary_contact = request.form.get('secondary_contact_number', '')
    secondary_contact_number = secondary_contact.strip() if secondary_contact else None
    
    home_address = request.form.get('home_address', '').strip()
    gender = request.form.get('gender')
    primary_id_type = request.form.get('primary_id_type')
    secondary_id_type = request.form.get('secondary_id_type')

    if not all([first_name, last_name, contact_number]):
        flash("Basic details (Name and Primary Contact) are required.", "error")
        return redirect(request.referrer or url_for('admin.customers'))

    existing = Customer.query.filter_by(
        first_name=first_name, 
        last_name=last_name, 
        contact_number=contact_number
    ).first()
    
    if existing:
        flash(f"A customer named {first_name} {last_name} with this contact number already exists.", "warning")
        return redirect(request.referrer or url_for('admin.customers'))

    upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'ids')
    os.makedirs(upload_folder, exist_ok=True)

    def save_id_file(file):
        if not file or file.filename == '':
            return None
        ext = os.path.splitext(file.filename)[1]
        filename = f"{uuid.uuid4().hex}{ext}"
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)
        return f"uploads/ids/{filename}"

    valid_id_file = request.files.get('valid_id')
    secondary_id_file = request.files.get('secondary_id')
    proof_of_billing_file = request.files.get('proof_of_billing')

    valid_id_path = save_id_file(valid_id_file)
    secondary_id_path = save_id_file(secondary_id_file)
    proof_of_billing_path = save_id_file(proof_of_billing_file)

    new_customer = Customer(
        user_id=None, 
        first_name=first_name,
        last_name=last_name,
        gender=gender,
        contact_number=contact_number,
        secondary_contact_number=secondary_contact_number,
        home_address=home_address,
        primary_id_type=primary_id_type,
        secondary_id_type=secondary_id_type,
        valid_id_path=valid_id_path,
        secondary_id_path=secondary_id_path,
        proof_of_billing_path=proof_of_billing_path, 
        id_uploaded_at=datetime.utcnow(),
        is_id_verified=True if valid_id_path else False,
        created_by_id=current_user.id,
        is_active=True 
    )

    try:
        db.session.add(new_customer)
        
        log = InventoryLog(
            action="Admin Created Customer",
            note=f"Customer {first_name} {last_name} added manually by {current_user.email}",
            user_id=current_user.id,
            user_name=f"{current_user.first_name} {current_user.last_name}"
        )
        db.session.add(log)
        
        db.session.commit()
        flash(f"Customer {new_customer.full_name} added successfully!", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Customer Creation Error: {str(e)}")
        flash("A database error occurred. Please try again.", "error")

    return redirect(url_for('admin.customers'))

@admin_bp.route('/update-customer', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_manage_customers')
def update_customer():
    customer_id = request.form.get("customer_id")
    customer = Customer.query.options(db.joinedload(Customer.user)).get_or_404(customer_id)

    first_name = request.form.get("first_name", "").strip().title()
    last_name = request.form.get("last_name", "").strip().title()
    
    if first_name:
        customer.first_name = first_name
        if customer.user: customer.user.first_name = first_name
    if last_name:
        customer.last_name = last_name
        if customer.user: customer.user.last_name = last_name

    gender = request.form.get("gender", "").strip().capitalize()
    if gender:
        customer.gender = gender

    customer.home_address = request.form.get("home_address", "").strip().title()
    customer.contact_number = request.form.get("contact_number", "").strip()
    customer.secondary_contact_number = request.form.get("secondary_contact_number", "").strip()

    if "is_active" in request.form:
        is_active_val = request.form.get("is_active") == "on"
        if customer.is_active != is_active_val:
            customer.is_active = is_active_val
            if customer.user:
                customer.user.is_active = is_active_val 

    def handle_id(file_key, path_attr, remove_flag_name=None, subfolder="ids", prefix="ID"):
        remove_flag = request.form.get(remove_flag_name) if remove_flag_name else None
        image_file = request.files.get(file_key)
        old_path = getattr(customer, path_attr, None)

        if remove_flag == "true" and old_path:
            full_path = os.path.join(current_app.root_path, 'static', old_path)
            if os.path.exists(full_path):
                os.remove(full_path)
            setattr(customer, path_attr, None)

        elif image_file and image_file.filename:
            ext = os.path.splitext(image_file.filename)[1].lower()
            filename = f"{prefix}_{uuid.uuid4().hex[:12]}{ext}"
            relative_path = f"uploads/{subfolder}/{filename}"
            absolute_path = os.path.join(current_app.root_path, 'static', 'uploads', subfolder)
            
            os.makedirs(absolute_path, exist_ok=True)
            
            if old_path:
                old_full = os.path.join(current_app.root_path, 'static', old_path)
                if os.path.exists(old_full):
                    os.remove(old_full)

            image_file.save(os.path.join(absolute_path, filename))
            setattr(customer, path_attr, relative_path)
            if file_key == "valid_id":
                customer.id_uploaded_at = datetime.utcnow()

    handle_id("valid_id", "valid_id_path", "remove_valid_id", "ids", "ID")
    handle_id("secondary_id", "secondary_id_path", "remove_secondary_id", "ids", "ID")
    handle_id("proof_of_billing", "proof_of_billing_path", None, "billing", "BILLING")

    customer.primary_id_type = request.form.get("primary_id_type", "").strip()
    customer.secondary_id_type = request.form.get("secondary_id_type", "").strip()

    if "is_id_verified" in request.form:
        customer.is_id_verified = request.form.get("is_id_verified") == "on"

    try:
        audit_log = InventoryLog(
            action="Admin Updated Customer",
            note=f"Customer ID {customer.id} updated by {current_user.first_name}",
            user_id=current_user.id,
            user_name=current_user.full_name
        )
        db.session.add(audit_log)
        
        db.session.commit()
        flash(f"Profile for {customer.full_name} updated successfully!", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Update Error: {e}")
        flash("An error occurred while saving changes.", "error")

    return redirect(url_for("admin.customers"))

@admin_bp.route('/products')
@limiter.exempt
@login_required
@admin_or_staff_required
@permission_required('can_manage_products')
def products():
    page           = request.args.get('page', 1, type=int)
    limit          = request.args.get('limit', 10, type=int)
    search_query   = request.args.get('q', '').strip()
    equipment_type = request.args.get('type', 'all')
    filter_type    = request.args.get('filter', '')         
 
    query = Product.query.filter(Product.status != 'Archived')
 
    if search_query:
        query = query.filter(or_(
            Product.equipment_type.ilike(f"%{search_query}%"),
            Product.name.ilike(f"%{search_query}%")
        ))
 
    if equipment_type and equipment_type != 'all':
        query = query.filter(Product.equipment_type.ilike(f"%{equipment_type}%"))
 
    # ── filter deep-links from notifications ──────────────────────────────
    if filter_type == 'low_stock':
        query = query.filter(
            Product.stock <= 5,
            Product.stock > 0,
            Product.status != 'Archived'
        )
    elif filter_type == 'out_of_stock':
        query = query.filter(
            Product.stock == 0,
            Product.status != 'Archived'
        )
    # ─────────────────────────────────────────────────────────────────────
 
    pagination = query.order_by(
        Product.equipment_type.asc(),
        Product.name.asc()
    ).paginate(page=page, per_page=limit, error_out=False)
 
    start_entry = 0
    end_entry   = 0
    if pagination.total > 0:
        start_entry = (pagination.page - 1) * pagination.per_page + 1
        end_entry   = min(pagination.page * pagination.per_page, pagination.total)
 
    stats = {
        'total_inventory':    db.session.query(func.sum(Product.stock)).scalar() or 0,
        'low_stock_count':    Product.query.filter(Product.stock <= 5).count(),
        'available_for_rent': Product.query.filter_by(status='Available').count()
    }
 
    customers = Customer.query.filter_by(is_active=True)\
        .order_by(Customer.last_name.asc()).all()
        
    # Dynamic equipment types from actual inventory
    equipment_types = db.session.query(
        Product.equipment_type
    ).filter(
        Product.equipment_type != None,
        Product.status != 'Archived'
    ).distinct().order_by(Product.equipment_type.asc()).all()
    equipment_types = [et[0] for et in equipment_types]
 
    return render_template(
        "admin/products.html",
        products=pagination.items,
        pagination=pagination,
        customers=customers,
        search_query=search_query,
        current_limit=limit,
        current_type=equipment_type,
        current_filter=filter_type,      
        start_entry=start_entry,
        end_entry=end_entry,
        equipment_types=equipment_types,
        **stats
    )

@admin_bp.route('/add-product', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_manage_products')
def add_product():
    equipment_type = request.form.get("equipment_type", "").strip().title()
    name = request.form.get("name", "").strip().title()
    size = request.form.get("size", "").strip()
    description = request.form.get("description", "").strip()
    
    is_refillable = True if request.form.get("is_refillable") == "on" else False

    if "oxygen" in equipment_type.lower() or "oxygen" in name.lower():
        is_refillable = True
    
    transaction_type = request.form.get("offer_type", "Sale").strip().title()
    rent_period = request.form.get("rent_period", "Monthly").strip().title()
    condition = request.form.get("condition", "Brand New").strip()

    try:
        stock = int(request.form.get("stock", 0))
        rent_price_raw = request.form.get("rent_price", "").strip()
        sale_price_raw = request.form.get("sale_price", "").strip()
        cost_price_raw = request.form.get("cost_price", "").strip()

        if transaction_type == 'Rental':
            rent_price = Decimal(rent_price_raw) if rent_price_raw else Decimal("0.00")
            sale_price = Decimal("0.00")
        elif transaction_type == 'Sale':
            sale_price = Decimal(sale_price_raw) if sale_price_raw else Decimal("0.00")
            rent_price = Decimal("0.00")
        else:
            rent_price = Decimal(rent_price_raw) if rent_price_raw else Decimal("0.00")
            sale_price = Decimal(sale_price_raw) if sale_price_raw else Decimal("0.00")

        cost_price = Decimal(cost_price_raw) if cost_price_raw else Decimal("0.00")
        
        if transaction_type == 'Rental' and rent_price <= 0:
            flash("Please provide a valid rent price for 'Rent Only' items.", "warning")
            return redirect(request.referrer)
        if transaction_type == 'Sale' and sale_price <= 0:
            flash("Please provide a valid sale price for 'Sale Only' items.", "warning")
            return redirect(request.referrer)
        
        if stock < 0:
            raise ValueError("Stock cannot be negative.")
        
    except (ValueError, InvalidOperation):
        flash("Oops! Please double-check the stock and price fields. Ensure you've entered valid numbers.", "error")
        return redirect(request.referrer)

    if not equipment_type:
        flash("The 'Equipment Type' is required. Please fill it out and try again.", "warning")
        return redirect(request.referrer)

    image_file = request.files.get("image")
    image_path = None
    
    if image_file and image_file.filename != '':
        ext = os.path.splitext(secure_filename(image_file.filename))[1].lower()
        if ext not in ['.jpg', '.jpeg', '.png', '.webp']:
            flash("Invalid image format. Please use JPG, PNG, or WebP files.", "error")
            return redirect(request.referrer)
            
        random_name = f"prod_{uuid.uuid4().hex[:12]}{ext}"
        upload_folder = os.path.join(current_app.root_path, "static", "uploads", "products")
        os.makedirs(upload_folder, exist_ok=True)
        
        image_path = f"uploads/products/{random_name}"
        full_path = os.path.join(upload_folder, random_name)
        
        try:
            image_file.save(full_path)
        except Exception as e:
            current_app.logger.error(f"Image Save Error: {e}")
            flash("We had trouble saving the product image. Please try uploading it again.", "error")
 
    try:
        new_product = Product(
            equipment_type=equipment_type,
            name=name,
            size=size, 
            description=description, 
            stock=stock,
            is_refillable=is_refillable,
            transaction_type=transaction_type,    
            rent_period=rent_period,    
            rent_price=rent_price,
            sale_price=sale_price,
            cost_price=cost_price, 
            condition=condition,   
            image=image_path,
            status="Available" if stock > 0 else "Out of Stock"
        )
        
        db.session.add(new_product)
        db.session.flush() 

        if is_refillable:
            new_tank_status = TankStatus(
                product_id=new_product.id,
                total_owned=stock,
                full_in_stock=stock, 
                empty_in_stock=0,
                rented_out=0
            )
            db.session.add(new_tank_status)

        log = InventoryLog(
            product_id=new_product.id,
            action="Initial Stock Entry",
            quantity=stock,
            note=f"Registered {name} (Size: {size}). Mode: {transaction_type}. Refillable: {is_refillable}. Condition: {condition}. Cost: {cost_price}",
            user_id=current_user.id,
            user_name=current_user.full_name 
        )
        db.session.add(log)

        if cost_price and cost_price > 0 and stock > 0:
            acq_category = get_cogs_category(new_product)
            acq_amount = (cost_price * stock).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            db.session.add(Expense(
                category=acq_category,
                expense_title=f"Stock Acquisition — {name} x{stock}",
                amount=acq_amount,
                description=f"Auto-recorded on initial stock entry. "
                            f"Unit cost: ₱{cost_price:,.2f} × {stock} unit(s).",
                date_incurred=datetime.utcnow().date(),
                recorded_by_id=current_user.id,
                product_id=new_product.id,
            ))
        elif stock > 0 and (not cost_price or cost_price == 0):
            db.session.add(InventoryLog(
                product_id=new_product.id,
                action="Acquisition Warning",
                note=f"Cost price not set for '{name}' — no acquisition expense recorded.",
                user_id=current_user.id,
                user_name=current_user.full_name
            ))
        
        db.session.commit()
        flash(f"Success! '{name}' has been added to your inventory.", "success")
        
    except Exception as e:
        db.session.rollback()
        if image_path:
            abs_image_path = os.path.join(current_app.root_path, "static", image_path)
            if os.path.exists(abs_image_path):
                os.remove(abs_image_path)
        
        current_app.logger.error(f"Database Error on Product Add: {str(e)}")
        flash("Something went wrong on our end while saving. Please wait a moment and try again.", "error")

    return redirect(url_for('admin.products'))

@admin_bp.route('/edit-product/<int:product_id>', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_manage_products')
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)

    new_type = request.form.get("equipment_type", "").strip().title()
    new_name = request.form.get("name", "").strip().title()
    new_size = request.form.get("size", "").strip() 
    new_description = request.form.get("description", "").strip()
    new_condition = request.form.get("condition", "").strip()
    
    new_offer_type = request.form.get("offer_type", "Rental").strip().title()
    new_rent_period = request.form.get("rent_period", "Monthly").strip().title()
    
    if not new_type or not new_name:
        flash("Equipment Type and Product Name are required fields.", "warning")
        return redirect(url_for('admin.products'))

    changes = []

    try:
        raw_rent = Decimal(request.form.get("rent_price") or "0.00")
        raw_sale = Decimal(request.form.get("sale_price") or "0.00")

        new_rent = raw_rent if new_offer_type in ['Both', 'Rental'] else Decimal("0.00")
        new_sale = raw_sale if new_offer_type in ['Both', 'Sale'] else Decimal("0.00")

        if product.transaction_type != new_offer_type:
            changes.append(f"Changed type from '{product.transaction_type}' to '{new_offer_type}'")
            product.transaction_type = new_offer_type

        if product.rent_period != new_rent_period:
            changes.append(f"Updated rental period to '{new_rent_period}'")
            product.rent_period = new_rent_period

        if product.equipment_type != new_type:
            changes.append(f"Updated type to '{new_type}'")
            product.equipment_type = new_type

        if (product.name or "") != new_name:
            changes.append(f"Renamed to '{new_name}'")
            product.name = new_name
            
        if (product.size or "") != new_size:
            changes.append(f"Updated size from '{product.size or 'N/A'}' to '{new_size}'")
            product.size = new_size

        if (product.description or "").strip() != new_description:
            changes.append("Updated description")
            product.description = new_description

        if product.condition != new_condition:
            changes.append(f"Changed condition to '{new_condition}'")
            product.condition = new_condition

        if product.rent_price != new_rent:
            changes.append(f"Changed rental rate to ₱{new_rent:,.2f}")
            product.rent_price = new_rent

        if product.sale_price != new_sale:
            changes.append(f"Changed sale price to ₱{new_sale:,.2f}")
            product.sale_price = new_sale

    except (ValueError, InvalidOperation):
        flash("Oops! Please double-check your pricing fields. Ensure you've entered valid numbers.", "error")
        return redirect(url_for('admin.products'))

    # Image Handling
    image_file = request.files.get("image")
    if image_file and image_file.filename != '':
        ext = os.path.splitext(secure_filename(image_file.filename))[1].lower()
        if ext not in ['.jpg', '.jpeg', '.png', '.webp']:
            flash("Invalid image format. Please use JPG, PNG, or WebP.", "error")
            return redirect(url_for('admin.products'))

        if product.image:
            old_full_path = os.path.join(current_app.root_path, "static", product.image)
            if os.path.exists(old_full_path):
                os.remove(old_full_path)

        random_name = f"prod_{uuid.uuid4().hex[:12]}{ext}"
        upload_folder = os.path.join(current_app.root_path, "static", "uploads", "products")
        os.makedirs(upload_folder, exist_ok=True)
        
        image_file.save(os.path.join(upload_folder, random_name))
        product.image = f"uploads/products/{random_name}"
        changes.append("Updated product image")

    try:
        if changes:
            inventory_log = InventoryLog(
                product_id=product.id,
                action="Product Edited",
                quantity=product.stock,
                note=f"{'; '.join(changes)}",
                user_id=current_user.id,
                user_name=current_user.full_name
            )
            db.session.add(inventory_log)
            db.session.commit()
            flash(f"Success! The details for '{product.name}' have been updated.", "success")
        else:
            flash("No changes were detected.", "info")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Product Update Error: {e}")
        flash("Something went wrong on our end while updating. Please try again.", "error")

    return redirect(url_for('admin.products'))

@admin_bp.route('/update_stock/<int:product_id>', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_manage_products')
@csrf.exempt
def update_stock(product_id):

    data = request.get_json() or {}
    product = Product.query.get_or_404(product_id)

    try:
        increment = int(data.get('increment', 0))
        reason = data.get('reason', '').strip() or 'Stock replenishment'
        total_amount_paid = data.get('total_amount_paid', None)
        if total_amount_paid is not None:
            total_amount_paid = Decimal(str(total_amount_paid)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
    except (ValueError, TypeError, Exception):
        return jsonify({
            "success": False,
            "message": "Invalid quantity or amount format."
        }), 400

    if increment <= 0:
        return jsonify({
            "success": False,
            "message": "Increment must be at least 1 unit."
        }), 400

    try:
        old_stock = product.stock
        product.stock += increment
        product.status = "Available" if product.stock > 0 else "Out of Stock"

        # ── Sync TankStatus for refillable products ────────────────────────
        if product.is_refillable and product.tank_status:
            product.tank_status.full_in_stock = (product.tank_status.full_in_stock or 0) + increment
            product.tank_status.total_owned = (
                (product.tank_status.full_in_stock or 0) +
                (product.tank_status.empty_in_stock or 0) +
                (product.tank_status.rented_out or 0)
            )

        # ── Auto-update unit cost if total amount paid is provided ─────────
        if total_amount_paid and total_amount_paid > 0:
            new_unit_cost = (total_amount_paid / increment).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            product.cost_price = new_unit_cost

        log_note = (
            f"Restocked: {old_stock} → {product.stock}. "
            f"Reason: {reason}"
        )
        if total_amount_paid:
            log_note += f". Total paid: ₱{total_amount_paid:,.2f}"

        db.session.add(InventoryLog(
            product_id=product.id,
            action="Restock",
            quantity=increment,
            note=log_note[:255],
            user_id=current_user.id,
            user_name=current_user.full_name
        ))

        # ── Auto restock expense ───────────────────────────────────────────
        expense_amount = total_amount_paid if (total_amount_paid and total_amount_paid > 0) else (
            (product.cost_price * increment).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if product.cost_price and product.cost_price > 0 else None
        )

        if expense_amount and expense_amount > 0:
            restock_category = get_cogs_category(product)
            unit_cost_display = product.cost_price or (total_amount_paid / increment)
            db.session.add(Expense(
                category=restock_category,
                expense_title=f"Restock — {product.name} x{increment}",
                amount=expense_amount,
                description=f"Auto-recorded on restock. "
                            f"Unit cost: ₱{unit_cost_display:,.2f} × {increment} unit(s).",
                date_incurred=datetime.utcnow().date(),
                recorded_by_id=current_user.id,
                product_id=product.id,
            ))
        else:
            db.session.add(InventoryLog(
                product_id=product.id,
                action="Acquisition Warning",
                note=f"No cost provided for restock of '{product.name}' — no expense recorded.",
                user_id=current_user.id,
                user_name=current_user.full_name
            ))

        db.session.commit()

        return jsonify({
            "success": True,
            "new_stock": product.stock,
            "new_status": product.status,
            "new_unit_cost": float(product.cost_price) if product.cost_price else 0,
            "message": (
                f"Added {increment} units. "
                f"Total stock for {product.name} is now {product.stock}."
                + (f" Unit cost updated to ₱{product.cost_price:,.2f}." if total_amount_paid else "")
            )
        })

    except Exception:
        db.session.rollback()
        current_app.logger.exception(f"Stock Update Error for Product {product_id}")
        return jsonify({
            "success": False,
            "message": "Unable to update inventory at this time."
        }), 500

     
@admin_bp.route('/delete-product/<int:product_id>', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_manage_products')
def delete_product(product_id):
    product = Product.query.get_or_404(product_id)

    try:
        has_active_rentals = Rental.query.filter(
            Rental.product_id == product.id,
            Rental.status.in_(["Active", "Overdue"])
        ).first()

        if has_active_rentals:
            flash(f"Cannot archive '{product.name}' due to active rentals.", "warning")
            return redirect(url_for('admin.products'))

        product.status = "Archived"
        product.stock = 0

        if hasattr(product, "is_active"):
            product.is_active = False

        if hasattr(product, "archived_at"):
            product.archived_at = datetime.utcnow()

        db.session.add(InventoryLog(
            product_id=product.id,
            action="ARCHIVE",
            quantity=0,
            note=f"Archived by {current_user.full_name}",
            user_id=current_user.id,
            user_name=current_user.full_name
        ))

        db.session.commit()

        flash(f"Product '{product.name}' archived successfully.", "success")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"ARCHIVE_ERROR: {str(e)}")
        flash("Failed to archive product.", "danger")

    return redirect(url_for('admin.products'))

@admin_bp.route('/process-purchase', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_process_transactions')
def process_purchase():
    from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
    
    is_ajax = request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    data = request.get_json() or request.form or {}

    def handle_response(success, message, redirect_url=None, status_code=200):
        """Helper to seamlessly manage both AJAX and traditional HTML Form responses."""
        if is_ajax:
            return jsonify({"success": success, "message": message}), status_code
        else:
            flash(message, "success" if success else "error")
            return redirect(redirect_url or url_for('admin.transactions'))

    try:
        try:
            amount_paid = Decimal(str(data.get('amount_paid', '0.00') or '0.00'))
        except (ValueError, TypeError, InvalidOperation):
            return handle_response(False, "Invalid numeric input for amount paid.", status_code=400)

        buyer_type = data.get('buyer_type', 'registered')
        customer_id = data.get('customer_id')

        customer_name = (data.get('unregistered_customer_name') or data.get('customer_name') or '').strip()

        if buyer_type == 'registered':
            if not customer_id:
                return handle_response(False, "Registered customer selection is required.", status_code=400)
            
            customer = db.session.get(Customer, int(customer_id))
            if not customer:
                return handle_response(False, "Selected customer profile not found.", status_code=404)

            customer_name = customer.full_name
        else:
            if not customer_name:
                return handle_response(False, "Walk-In Buyer / Patient full name is required.", status_code=400)
            customer_id = None

        items_data = data.get('items', [])
        if isinstance(items_data, str):
            try:
                items_data = json.loads(items_data)
            except (json.JSONDecodeError, TypeError):
                return handle_response(False, "Failed to decode basket collection elements.", status_code=400)

        if not items_data or not isinstance(items_data, list):
            return handle_response(False, "Please select at least one item to purchase.", status_code=400)

        now = datetime.now()
        ref_no = f"PUR-{now:%m%d%Y}-{now:%H%M%S}-{''.join(random.choices(string.ascii_uppercase + string.digits, k=4))}"
        fulfillment = data.get("fulfillment_type", "Walk-In")

        new_transaction = Transaction(
            reference_no=ref_no,
            customer_id=customer_id,
            customer_name=customer_name,
            processed_by=current_user.id,
            transaction_type="Sale",
            total_amount=Decimal("0.00"),
            fulfillment_type=fulfillment,
            delivery_address=data.get("delivery_address") if fulfillment == "Delivery" else None,
            landmark=data.get("landmark") if fulfillment == "Delivery" else None,
            delivery_status="Pending" if fulfillment == "Delivery" else "N/A",
            status="Open"
        )

        db.session.add(new_transaction)
        db.session.flush() 

        total_transaction_price = Decimal('0.00')

        for item in items_data:
            try:
                product_id = int(item.get('id'))
                quantity = int(item.get('quantity', 1))
                unit_price = Decimal(str(item.get('price', '0.00')))
            except (ValueError, TypeError, InvalidOperation):
                db.session.rollback()
                return handle_response(False, "Invalid numeric values detected inside basket components.", status_code=400)

            if quantity <= 0:
                db.session.rollback()
                return handle_response(False, "Quantity must be at least 1 for all selected products.", status_code=400)

            product = Product.query.filter_by(id=product_id).with_for_update().first()

            if not product:
                db.session.rollback()
                return handle_response(False, f"Product entry unique ID {product_id} is unavailable.", status_code=404)

            if not product.is_active:
                db.session.rollback()
                return handle_response(False, f"'{product.name}' has been archived.", status_code=400)

            if product.stock < quantity:
                db.session.rollback()
                return handle_response(False, f"Only {product.stock} units left in stock for '{product.name}'.", status_code=400)

            item_total_price = (unit_price * quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            total_transaction_price += item_total_price

            product.stock -= quantity

            if product.stock <= 0:
                product.stock = 0
                product.status = "Out of Stock"
            else:
                product.status = "Available"
                
            # ── Sync TankStatus for refillable products on sale ───────────
            if product.is_refillable and product.tank_status:
                product.tank_status.full_in_stock = max(
                    (product.tank_status.full_in_stock or 0) - quantity, 0
                )
                product.tank_status.total_owned = max(
                    (product.tank_status.total_owned or 0) - quantity, 0
                )

            db.session.add(Purchase(
                transaction_id=new_transaction.id,
                product_id=product.id,
                customer_id=customer_id,  
                quantity=quantity,
                unit_price=unit_price,
                total_price=item_total_price,
                warranty_or_notes=(data.get("warranty_or_notes") or "").strip(),
                product_name=product.name,
                product_asset_tag=product.asset_tag,
                product_type=getattr(product, 'equipment_type', None) or getattr(product, 'product_type', None),
                product_description=product.description,
                product_image=product.image,
                product_condition=product.condition,
                product_category=product.category,
                product_cost_price=product.cost_price,
                product_sale_price=product.sale_price,
                product_rent_price=product.rent_price,
                product_rent_period=product.rent_period
            ))

            db.session.add(InventoryLog(
                product_id=product.id,
                action="SALE",
                quantity=-quantity,
                note=f"Sold {quantity} via {ref_no}",
                user_id=current_user.id,
                user_name=getattr(current_user, 'full_name', 'System Administrator')
            ))



        new_transaction.total_amount = total_transaction_price

        if amount_paid > 0:

            payment_method = (data.get("payment_method") or "Cash").strip()

            reference_number = ((data.get("reference_number") or "").strip())


            if payment_method.lower() == "cash":
                reference_number = None

            else:

                if not reference_number:
                    db.session.rollback()
                    return handle_response(
                        False,
                        "Reference number is required for non-cash payments.",
                        status_code=400
                    )

                existing_payment = Payment.query.filter_by(
                    reference_number=reference_number
                ).first()

                if existing_payment:
                    db.session.rollback()
                    return handle_response(
                        False,
                        "Payment reference number already exists.",
                        status_code=400
                    )

            db.session.add(Payment(
                transaction_id=new_transaction.id,
                amount=amount_paid,
                payment_method=payment_method,
                reference_number=reference_number,
                status="Completed",
                verified_by_id=current_user.id,
                verified_at=datetime.utcnow()
            ))

        if hasattr(new_transaction, 'update_totals'):
            new_transaction.update_totals()
            
        db.session.commit()
        return handle_response(True, f"Purchase {ref_no} completed successfully!")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"PURCHASE_ERROR: {str(e)}")
        return handle_response(False, f"Server error occurred during purchase processing sequence.", status_code=500)
       
       
@admin_bp.route('/process-rental', methods=['POST'])
@login_required
@admin_or_staff_required
@permission_required('can_process_transactions')
def process_rental():
    try:
        product_ids = request.form.getlist('product_id[]') or request.form.getlist('product_id')
        quantities = request.form.getlist('quantity[]') or request.form.getlist('quantity')
        unit_prices = request.form.getlist('unit_price[]') or request.form.getlist('unit_price')

        serial_number = request.form.get('serial_number')

        if not product_ids:
            flash("No products selected.", "danger")
            return redirect(request.referrer)

        fulfillment_type = request.form.get('fulfillment_type', 'Walk-In')
        amount_paid = Decimal(request.form.get('amount_paid', '0').replace(',', '') or '0')

        start_date = datetime.strptime(request.form.get('start_date'), "%Y-%m-%d").date()
        expected_return = datetime.strptime(request.form.get('return_date'), "%Y-%m-%d").date()

        if expected_return < start_date:
            flash("Invalid return date.", "danger")
            return redirect(request.referrer)

        ref_no = f"RNT-{datetime.now():%m%d%Y-%H%M%S}-{''.join(random.choices(string.ascii_uppercase+string.digits, k=4))}"

        new_txn = Transaction(
            reference_no=ref_no,
            customer_id=request.form.get('customer_id'),
            transaction_type="Rental",
            total_amount=Decimal("0.00"),
            fulfillment_type=fulfillment_type,
            status="Open"
        )

        db.session.add(new_txn)
        db.session.flush()

        rentals = []

        for i, pid in enumerate(product_ids):
            product = Product.query.get_or_404(pid)
            qty = int(quantities[i])
            price = Decimal(unit_prices[i])

            if qty <= 0:
                db.session.rollback()
                flash(f"Invalid quantity for {product.name}", "danger")
                return redirect(request.referrer)

            if product.is_refillable:
                if not product.tank_status:
                    db.session.rollback()
                    flash(f"Tank status not configured for {product.name}", "danger")
                    return redirect(request.referrer)

                available = product.tank_status.full_in_stock or 0
                if available < qty:
                    db.session.rollback()
                    flash(f"Not enough available tanks for {product.name}. Available: {available}", "danger")
                    return redirect(request.referrer)

                product.tank_status.rented_out = (product.tank_status.rented_out or 0) + qty
                product.tank_status.full_in_stock = max((product.tank_status.full_in_stock or 0) - qty, 0)
                log_note = f"Rental created (Tank metrics updated) {ref_no}"
            else:
                if hasattr(product, 'stock') and product.stock is not None:
                    if product.stock < qty:
                        db.session.rollback()
                        flash(f"Not enough stock available for {product.name}. Available: {product.stock}", "danger")
                        return redirect(request.referrer)
                    
                    product.stock -= qty
                log_note = f"Rental created (Standard asset stock deducted) {ref_no}"

            rental = Rental(
                transaction_id=new_txn.id,
                product_id=product.id,
                customer_id=new_txn.customer_id,
                start_date=start_date,
                expected_return_date=expected_return,
                monthly_rate=price,
                quantity=qty,
                serial_number=serial_number,
                status="Active"
            )

            db.session.add(rental)
            rentals.append(rental)

            db.session.add(InventoryLog(
                product_id=product.id,
                action="RENTAL",
                quantity=qty if not product.is_refillable else 0,  
                note=log_note,
                user_id=current_user.id,
                user_name=current_user.full_name
            ))

        db.session.flush()

        for r in rentals:
            r.generate_monthly_invoices()

        db.session.flush()
        db.session.expire(new_txn, ['payments'])
        
        # ── Record initial payment if provided ────────────────────────────
        amount_paid_raw = request.form.get('amount_paid', '0').replace(',', '') or '0'
        amount_paid = Decimal(amount_paid_raw)

        if amount_paid > 0:
            payment_method = request.form.get('payment_method', 'Cash').strip()
            reference_number = request.form.get('reference_number', '').strip() or None

            # Distribute initial payment across first unpaid invoice(s)
            remaining = amount_paid
            all_invoices = []
            for r in rentals:
                for inv in r.invoices:
                    all_invoices.append(inv)

            for inv in all_invoices:
                if remaining <= 0:
                    break
                pay_amount = min(remaining, Decimal(str(inv.amount_due or 0)))
                db.session.add(Payment(
                    transaction_id=new_txn.id,
                    invoice_id=inv.id,
                    amount=pay_amount,
                    payment_method=payment_method,
                    reference_number=reference_number,
                    status="Completed",
                    verified_by_id=current_user.id,
                    verified_at=datetime.utcnow()
                ))
                if pay_amount >= Decimal(str(inv.amount_due or 0)):
                    inv.status = "Paid"
                else:
                    inv.status = "Partially Paid"
                remaining -= pay_amount

        db.session.flush()
        db.session.expire(new_txn, ['payments'])
        new_txn.update_totals()
        db.session.commit()

        flash(f"Rental created: {ref_no}", "success")
        return redirect(url_for("admin.transactions"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"RENTAL_ERROR: {str(e)}")
        flash("Rental processing failed.", "danger")
        return redirect(request.referrer)
    
     
@admin_bp.route('/product/<int:product_id>/history')
@login_required
@admin_or_staff_required
def product_history(product_id):
    try:
        logs = InventoryLog.query.filter_by(product_id=product_id)\
            .order_by(InventoryLog.created_at.desc())\
            .limit(100)\
            .all()

        if not logs:
            product_exists = Product.query.get(product_id)
            if not product_exists:
                return jsonify({"status": "error", "message": "Product not found"}), 404
            return jsonify([]) 

        result = []
        for log in logs:
            movement_type = "neutral"
            if log.quantity and log.quantity > 0:
                movement_type = "increase"
            elif log.quantity and log.quantity < 0:
                movement_type = "decrease"

            result.append({
                "id": log.id,
                "action": log.action,
                "quantity": log.quantity if log.quantity != 0 else "-",
                "movement_type": movement_type, 
                "note": log.note or "No additional notes.",
                "user": log.user_name or "System",
                "date": log.created_at.strftime("%b %d, %Y"),
                "time": log.created_at.strftime("%I:%M %p"),
                "timestamp": log.created_at.isoformat() 
            })

        return jsonify(result)

    except Exception as e:
        current_app.logger.error(f"Error fetching history for Product {product_id}: {str(e)}")
        return jsonify({
            "status": "error", 
            "message": "Internal server error while fetching history."
        }), 500


@admin_bp.route('/transactions')
@login_required
@admin_or_staff_required
@csrf.exempt
@permission_required('can_process_transactions')
def transactions():
    current_date = datetime.now().date()

    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 10, type=int)
    search_query = request.args.get('q', '').strip()
    txn_type = request.args.get('type', '')
    fulfillment = request.args.get('fulfillment', '')
    status_filter = request.args.get('status', '')

    query = Transaction.query.options(
        joinedload(Transaction.customer),
        selectinload(Transaction.rentals).selectinload(Rental.invoices),
        selectinload(Transaction.payments)
    )

    if search_query:
        query = query.outerjoin(Rental, Rental.transaction_id == Transaction.id).filter(or_(
            Transaction.reference_no.ilike(f"%{search_query}%"),
            Transaction.customer_name.ilike(f"%{search_query}%"),
            Transaction.landmark.ilike(f"%{search_query}%"),
            Rental.serial_number.ilike(f"%{search_query}%")
        )).distinct()
    
    if txn_type:
        query = query.filter(Transaction.transaction_type == txn_type)
    
    if fulfillment:
        query = query.filter(Transaction.fulfillment_type == fulfillment)
 
    elif status_filter == 'expiring':
        from datetime import timedelta
        soon = current_date + timedelta(days=5)
        query = query.filter(
            Transaction.transaction_type == 'Rental',
            Transaction.rentals.any(
                and_(
                    Rental.status == 'Active',
                    Rental.expected_return_date >= current_date,
                    Rental.expected_return_date <= soon
                )
            )
        )
        
    if status_filter == 'overdue_return':
        # Rentals NOT RETURNED past due date
        query = query.filter(
            Transaction.transaction_type == 'Rental',
            Transaction.rentals.any(
                and_(
                    Rental.status == 'Active',
                    Rental.expected_return_date < current_date
                )
            )
        )

    elif status_filter == 'overdue_payment':
        # Rental payments past due
        query = query.filter(
            Transaction.transaction_type == 'Rental',
            Transaction.balance_due > 0,
            Transaction.status == 'Open'
        )
 
    elif status_filter == 'unpaid':
        query = query.filter(
            Transaction.payment_status == 'Unpaid',
            Transaction.status == 'Open'
        )
 
    elif status_filter == 'partial':
        query = query.filter(
            Transaction.payment_status == 'Partially Paid',
            Transaction.status == 'Open'
        )
 
    elif status_filter == 'submitted':
        query = query.filter(
            Transaction.tracking_status == 'SUBMITTED',
            Transaction.status == 'Open'
        )

    pagination = query.order_by(Transaction.created_at.desc()).paginate(
        page=page, per_page=limit, error_out=False
    )

    for txn in pagination.items:
        rental_overdue = False
        if txn.rentals and txn.rentals[0].expected_return_date:
            ret_date = txn.rentals[0].expected_return_date
            if hasattr(ret_date, 'date'):
                ret_date = ret_date.date()
            rental_overdue = ret_date < current_date
            
        txn.calculated_overdue = txn.balance_due > 0 and (getattr(txn, 'is_overdue', False) or rental_overdue)

    stats = {
        'pending': Transaction.query.filter_by(payment_status='Unpaid').count(),
        'alerts': Transaction.query.filter(Transaction.status == 'Due').count(),
        'empty_tanks': Product.query.filter(Product.status.ilike('%Empty%')).count()
    }
    
    customers = Customer.query.order_by(Customer.last_name).all()
    all_equipment = Product.query.filter_by(status='Available').order_by(Product.equipment_type.asc()).all()
    
    for equipment in all_equipment:
        if equipment.is_refillable and equipment.tank_status:
            equipment.available_stock = equipment.tank_status.full_in_stock or 0
        else:
            equipment.available_stock = equipment.stock or 0

    raw_refillables = Product.query.join(TankStatus).filter(Product.is_active == True).all()
    grouped_refills = {}
    for p in raw_refillables:
        key = (p.name, p.size)
        if key not in grouped_refills:
            grouped_refills[key] = {'name': p.name, 'size': p.size}
    
    unique_refillable_products = list(grouped_refills.values())

    return render_template(
        "admin/transactions.html",
        transactions=pagination.items,
        pagination=pagination,
        search_query=search_query,
        current_limit=limit,
        current_type=txn_type,
        current_fulfillment=fulfillment,
        current_status=status_filter, 
        customers=customers,
        all_equipment=all_equipment,
        refillable_products=unique_refillable_products,
        datetime_now_date=current_date,
        **stats
    )
    
@admin_bp.route('/active-rentals')
@login_required
@admin_or_staff_required
@permission_required('can_view_active_rentals')
def active_rentals():
    from datetime import date

    search_query = request.args.get('q', '').strip()
    filter_type = request.args.get('filter', '')  
    
    query = Rental.query.filter_by(status='Active')\
        .options(
            joinedload(Rental.product),
            joinedload(Rental.transaction).joinedload(Transaction.customer)
        )

    today = date.today()
    if filter_type == 'overdue_return':
        query = query.filter(Rental.expected_return_date < today)
    
    rentals = query.order_by(Rental.expected_return_date.asc()).all()
    
    return render_template(
        'admin/active_rentals.html',
        rentals=rentals,
        search_query=search_query,
        datetime_now_date=date.today(),
        current_filter=filter_type 
    )

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from flask import make_response

@admin_bp.route('/transactions/export')
@login_required
@admin_or_staff_required
@permission_required('can_process_transactions')
def export_transactions():
    search_query = request.args.get('q', '').strip()
    txn_type = request.args.get('type', '')
    fulfillment = request.args.get('fulfillment', '')
    status_filter = request.args.get('status', '')

    query = Transaction.query.options(
        joinedload(Transaction.customer),
        selectinload(Transaction.rentals)
    )

    if search_query:
        query = query.filter(or_(
            Transaction.reference_no.ilike(f"%{search_query}%"),
            Transaction.customer_name.ilike(f"%{search_query}%"),
            Transaction.landmark.ilike(f"%{search_query}%")
        ))
    if txn_type:
        query = query.filter(Transaction.transaction_type == txn_type)
    if fulfillment:
        query = query.filter(Transaction.fulfillment_type == fulfillment)
    if status_filter == 'overdue':
        current_date = datetime.now().date()
        query = query.filter(
            Transaction.balance_due > 0,
            or_(
                Transaction.status == 'Due', 
                Transaction.rentals.any(Rental.expected_return_date < current_date)
            )
        )

    transactions_list = query.order_by(Transaction.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid") # Matching dashboard green
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = ['Reference No', 'Customer Name', 'Type', 'Fulfillment', 'Total Amount', 'Amount Paid', 'Balance Due', 'Status', 'Date']
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for txn in transactions_list:
        cust_name = txn.customer_name or (f"{txn.customer.first_name} {txn.customer.last_name}" if txn.customer else "—")
        txn_date = txn.created_at.strftime('%Y-%m-%d') if txn.created_at else "—"
        
        row_data = [
            txn.reference_no,
            cust_name,
            txn.transaction_type,
            txn.fulfillment_type,
            txn.total_amount,
            txn.amount_paid,
            txn.balance_due,
            txn.status,
            txn_date
        ]
        ws.append(row_data)
        
        current_row = ws.max_row
        ws.cell(row=current_row, column=5).number_format = '"₱"#,##0.00' 
        ws.cell(row=current_row, column=6).number_format = '"₱"#,##0.00' 
        ws.cell(row=current_row, column=7).number_format = '"₱"#,##0.00' 
        
        ws.cell(row=current_row, column=5).alignment = right_align
        ws.cell(row=current_row, column=6).alignment = right_align
        ws.cell(row=current_row, column=7).alignment = right_align
        ws.cell(row=current_row, column=8).alignment = center_align
        ws.cell(row=current_row, column=9).alignment = center_align

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send response file download stream
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=transactions_export.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@admin_bp.route('/transaction_details/<int:id>') 
@login_required
@admin_or_staff_required
def transaction_details(id):
    txn = Transaction.query.get_or_404(id)
    
    return render_template('admin/transaction_details.html', txn=txn, current_date=datetime.now().date())

@admin_bp.route('/transaction/<int:txn_id>/update-tracking', methods=['POST'])
@login_required
def update_tracking(txn_id):
    txn = Transaction.query.get_or_404(txn_id)
    new_status = request.form.get('tracking_status')

    valid_statuses = ['SUBMITTED', 'VERIFIED', 'SHIPPED', 'DELIVERED', 'CANCELLED']

    if new_status in valid_statuses:
        if txn.tracking_status == 'CANCELLED':
            flash("Action denied: This transaction has already been canceled.", "warning")
            return redirect(url_for('admin.transaction_details', id=txn.id))

        txn.tracking_status = new_status

        if new_status == 'CANCELLED':
            txn.status = 'Cancelled' 
        elif new_status == 'DELIVERED':
            txn.delivery_status = 'Delivered'
            if txn.transaction_type != 'Rental':
                txn.status = 'Closed'

        try:
            txn.update_totals()
            db.session.commit()
            flash(f"Success: Status updated to '{new_status.title()}'.", "success")
        except Exception as e:
            db.session.rollback()
            flash("System Error: Could not update status.", "danger")
    else:
        flash("Update Failed: Invalid status provided.", "danger")

    return redirect(url_for('admin.transaction_details', id=txn.id))

@admin_bp.route('/post-payment', methods=['POST'])
@login_required
@admin_or_staff_required
def post_payment():
    txn_id = request.form.get('txn_id')
    invoice_id = request.form.get('invoice_id')  

    try:
        raw_amount = request.form.get('amount', '0').replace(',', '').strip()
        amount = Decimal(raw_amount)

        if amount <= 0:
            flash('Payment amount must be greater than zero.', 'warning')
            return redirect(request.referrer)

        txn = Transaction.query.with_for_update().get_or_404(txn_id)

        allowed_methods = ["Cash", "GCash", "Bank Transfer", "Check"]
        method = request.form.get('method')

        if method not in allowed_methods:
            flash("Invalid payment method.", "danger")
            return redirect(request.referrer)

        ref_number = request.form.get('payment_reference', '').strip() or None

        if method in ["GCash", "Bank Transfer", "Check"] and not ref_number:
            flash("Reference number is required for this payment method.", "warning")
            return redirect(request.referrer)

        if ref_number and Payment.query.filter_by(reference_number=ref_number).first():
            flash("Reference number already exists.", "danger")
            return redirect(request.referrer)

        receipt_path = None
        file = request.files.get('receipt_image')
        if file and file.filename:
            filename = secure_filename(f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}")
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'receipts')
            os.makedirs(upload_folder, exist_ok=True)
            file.save(os.path.join(upload_folder, filename))
            receipt_path = f'uploads/receipts/{filename}'

        if txn.transaction_type == 'Sale':
            sale_balance = Decimal(str(txn.balance_due or 0))
            if sale_balance <= 0:
                flash('This transaction is already fully paid.', 'info')
                return redirect(request.referrer)
            if amount > sale_balance:
                flash(f'Payment ₱{amount:,.2f} exceeds the remaining balance of ₱{sale_balance:,.2f}.', 'danger')
                return redirect(request.referrer)

            payment = Payment(
                transaction_id=txn.id, 
                amount=amount, 
                payment_method=method, 
                reference_number=ref_number, 
                receipt_image_path=receipt_path,
                status="Completed", 
                verified_by_id=current_user.id, 
                verified_at=datetime.utcnow()
            )
            db.session.add(payment)

        else:
            if not invoice_id:
                flash("Please select a specific item/invoice to pay.", "warning")
                return redirect(request.referrer)

            rental_ids = [r.id for r in txn.rentals]
            target_invoice = RentalInvoice.query.filter(
                RentalInvoice.id == invoice_id,
                RentalInvoice.rental_id.in_(rental_ids)
            ).first()

            if not target_invoice:
                flash("Invalid invoice selected.", "danger")
                return redirect(request.referrer)

            if amount > target_invoice.remaining_balance:
                flash(f'Payment ₱{amount:,.2f} exceeds the invoice balance of ₱{target_invoice.remaining_balance:,.2f}.', 'danger')
                return redirect(request.referrer)

            payment = Payment(
                transaction_id=txn.id, 
                invoice_id=target_invoice.id, 
                amount=amount,
                payment_method=method, 
                reference_number=ref_number, 
                receipt_image_path=receipt_path, 
                status="Completed",
                verified_by_id=current_user.id, 
                verified_at=datetime.utcnow()
            )
            db.session.add(payment)

            if amount >= target_invoice.remaining_balance:
                target_invoice.status = "Paid"
            else:
                target_invoice.status = "Partially Paid"

        db.session.expire(txn, ['payments'])
        txn.update_totals()
        db.session.commit()

        flash(f'Payment of ₱{amount:,.2f} recorded successfully.', 'success')

    except (InvalidOperation, ValueError):
        db.session.rollback()
        flash('Invalid amount format.', 'danger')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"PAYMENT_ERROR | TXN: {txn_id} | Error: {str(e)}")
        flash(f'A system error occurred: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('admin.transactions'))

@admin_bp.route('/confirm-payment-proof/<int:proof_id>', methods=['POST'])
@login_required
@admin_or_staff_required
def confirm_payment_proof(proof_id):
    proof = PaymentProof.query.get_or_404(proof_id)
    txn = Transaction.query.get_or_404(proof.transaction_id)

    try:
        # Guard: don't double-record the same reference number
        existing_payment = Payment.query.filter_by(
            reference_number=proof.reference_number
        ).first()

        if existing_payment:
            # Already recorded — just mark the proof as verified
            proof.status = 'Verified'
            proof.payment_id = existing_payment.id
            db.session.flush()
            db.session.expire(txn, ['payments'])
            txn.update_totals()
            db.session.commit()
            flash('Payment proof verified successfully.', 'success')
            return redirect(url_for('admin.transaction_details', id=txn.id))

        # No existing payment — create one for the remaining balance
        remaining = Decimal(str(txn.balance_due or 0))

        if remaining <= 0:
            flash('This transaction is already fully paid.', 'info')
            return redirect(url_for('admin.transaction_details', id=txn.id))

        # For rentals: attach to the oldest unpaid invoice
        invoice_id_to_use = None
        if txn.transaction_type == 'Rental':
            unpaid_invoice = RentalInvoice.query.filter(
                RentalInvoice.rental_id.in_([r.id for r in txn.rentals]),
                RentalInvoice.status != 'Paid'
            ).order_by(RentalInvoice.service_period_start.asc()).first()

            if unpaid_invoice:
                invoice_id_to_use = unpaid_invoice.id
                pay_amount = min(remaining, unpaid_invoice.remaining_balance)

                if pay_amount >= unpaid_invoice.remaining_balance:
                    unpaid_invoice.status = 'Paid'
                else:
                    unpaid_invoice.status = 'Partially Paid'
            else:
                pay_amount = remaining
        else:
            pay_amount = remaining

        new_payment = Payment(
            transaction_id=txn.id,
            invoice_id=invoice_id_to_use,
            amount=pay_amount,
            payment_method='GCash',
            reference_number=proof.reference_number,
            status='Completed',
            verified_by_id=current_user.id,
            verified_at=datetime.utcnow()
        )
        db.session.add(new_payment)
        db.session.flush()

        proof.status = 'Verified'
        proof.payment_id = new_payment.id

        # ── Let update_totals() handle all status strings ──
        db.session.expire(txn, ['payments'])
        txn.update_totals()
        db.session.commit()

        flash('Payment confirmed and marked as Paid successfully.', 'success')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"PROOF_CONFIRM_ERROR | proof {proof_id} | {str(e)}")
        flash(f'Failed to confirm payment: {str(e)}', 'error')

    return redirect(url_for('admin.transaction_details', id=txn.id))


@admin_bp.route('/transaction/<int:txn_id>/return', methods=['POST'])
@login_required
@admin_or_staff_required
def process_return(txn_id):
    txn = Transaction.query.get_or_404(txn_id)
    returned_item_ids = request.form.getlist('returned_items')
    return_notes = request.form.get('return_notes', '').strip()
    raw_late_fees = request.form.get('late_fees', '0')

    if not returned_item_ids:
        flash("No items were selected to return.", "warning")
        return redirect(url_for('admin.transaction_details', id=txn.id))

    try:
        late_fees = Decimal(str(raw_late_fees))
    except (InvalidOperation, ValueError, TypeError):
        late_fees = Decimal('0.00')

    try:
        user_display_name = f"{current_user.first_name} {current_user.last_name}"
        items_processed = 0

        for item_id in returned_item_ids:
            rental = Rental.query.get(item_id)
            if not rental or rental.transaction_id != txn.id:
                continue

            qty_input = request.form.get(f'qty_{item_id}', '0')
            qty_to_return = int(qty_input) if qty_input.isdigit() else 0
            
            if qty_to_return > rental.remaining_to_return:
                qty_to_return = rental.remaining_to_return
            
            if qty_to_return > 0:
                rental.quantity_returned = (rental.quantity_returned or 0) + qty_to_return
                if rental.quantity_returned >= rental.quantity:
                    rental.status = 'Returned'
                    rental.actual_return_date = datetime.utcnow().date()
                else:
                    rental.status = 'Partially Returned'
                
                return_status = request.form.get(f'status_{item_id}', 'Full')
                
                # --- REFILLABLE TANK RETURN LOGIC ---
                if rental.product.tank_status:
                    tank_info = rental.product.tank_status
                    
                    # 1. Bring back the tank from the field
                    tank_info.rented_out = max(0, (tank_info.rented_out or 0) - qty_to_return)
                    
                    # 2. Sort into either empty or full stock tiers
                    if return_status == 'Empty':
                        tank_info.empty_in_stock = (tank_info.empty_in_stock or 0) + qty_to_return
                        # Note: core product.stock is NOT increased because an empty tank can't be rented yet
                    else:
                        tank_info.full_in_stock = (tank_info.full_in_stock or 0) + qty_to_return
                        # Core product.stock increases because a full tank is ready to go out again
                        rental.product.stock = (rental.product.stock or 0) + qty_to_return
                    
                    # Clean up product availability status dynamically
                    if rental.product.stock > 0:
                        rental.product.status = "Available"
                        
                    db.session.add(tank_info)
                    db.session.add(rental.product)
                
                # --- STANDARD NON-REFILLABLE EQUIPMENT RETURN LOGIC ---
                else:
                    rental.product.stock = (rental.product.stock or 0) + qty_to_return
                    if rental.product.stock > 0:
                        rental.product.status = "Available"
                    db.session.add(rental.product)

                db.session.add(InventoryLog(
                    product_id=rental.product.id,
                    action="Equipment Return",
                    quantity=qty_to_return,
                    note=f"Txn {txn.reference_no} Return. Status: {return_status}. {return_notes}",
                    user_id=current_user.id,
                    user_name=user_display_name
                ))
                items_processed += 1

        if late_fees > 0:
            txn.total_late_fees = (Decimal(str(txn.total_late_fees or 0)) + late_fees)
        
        if all(r.status == 'Returned' for r in txn.rentals):
            txn.status = 'Closed'

        db.session.commit()
        flash(f"Successfully processed return for {items_processed} item(s).", "success")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"RETURN_ERROR | TXN: {txn_id} | {str(e)}")
        flash("An error occurred while saving the return.", "danger")

    return redirect(url_for('admin.transaction_details', id=txn.id))


@admin_bp.route('/process-primegas', methods=['POST'])
@login_required
def process_primegas():
    product_id = request.form.get('product_id')
    quantity = int(request.form.get('quantity', 0))

    tank_status = TankStatus.query.filter_by(product_id=product_id).first()

    if not tank_status or tank_status.empty_in_stock < quantity:
        flash('Insufficient empty stock.', 'danger')
        return redirect(url_for('admin.transactions'))

    tank_status.empty_in_stock -= quantity
    tank_status.full_in_stock += quantity
    
    db.session.commit()
    flash('Inventory updated successfully!', 'success')
    return redirect(url_for('admin.transactions'))

@admin_bp.route('/system_logs')
@limiter.exempt
@login_required
@admin_or_staff_required
@permission_required('can_view_reports')
def system_logs():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 10, type=int)
    search_query = request.args.get('q', '')
    current_type = request.args.get('type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = InventoryLog.query

    if search_query:
        from sqlalchemy import func
        query = query.filter(or_(
            InventoryLog.action.ilike(f'%{search_query}%'),
            InventoryLog.user_name.ilike(f'%{search_query}%'),
            InventoryLog.note.ilike(f'%{search_query}%'),
            func.date_format(InventoryLog.created_at, '%b %d, %Y').ilike(f'%{search_query}%'),
            func.date_format(InventoryLog.created_at, '%M %d, %Y').ilike(f'%{search_query}%'),
            func.date_format(InventoryLog.created_at, '%b %Y').ilike(f'%{search_query}%'),
        ))

    if current_type:
        query = query.filter(InventoryLog.action == current_type)

    if date_from:
        try:
            from datetime import datetime
            date_from_dt = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(InventoryLog.created_at >= date_from_dt)
        except ValueError:
            pass

    if date_to:
        try:
            from datetime import datetime
            date_to_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(InventoryLog.created_at <= date_to_dt)
        except ValueError:
            pass

    staff_filter = request.args.get('staff_id', '', type=str)
    if staff_filter:
        try:
            query = query.filter(InventoryLog.user_id == int(staff_filter))
        except (ValueError, TypeError):
            pass

    pagination = query.order_by(InventoryLog.created_at.desc()).paginate(
        page=page,
        per_page=limit,
        error_out=False
    )
    
    # Staff list for filter dropdown
    staff_list = User.query.filter(
        User.role.in_(['Staff', 'Administrator'])
    ).order_by(User.first_name).all()

    return render_template(
        'admin/system_logs.html',
        all_logs=pagination.items,
        pagination=pagination,
        current_limit=limit,
        search_query=search_query,
        current_type=current_type,
        date_from=date_from,
        date_to=date_to,
        staff_list=staff_list,
        staff_filter=staff_filter,
    )

@admin_bp.route('/logs/<int:log_id>/detail')
@login_required
@admin_or_staff_required
def log_detail(log_id):
    log = InventoryLog.query.get_or_404(log_id)
    return jsonify({
        'status':   'success',
        'action':   log.action,
        'user':     log.user_name or 'System',
        'product':  log.product.name if log.product else 'N/A',
        'asset':    log.product.asset_tag if log.product else '—',
        'quantity': log.quantity or 0,
        'note':     log.note or 'No notes.',
        'date':     log.created_at.strftime('%b %d, %Y %I:%M %p') if log.created_at else '—',
    })

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'webp'}

@admin_bp.route('/profile', methods=['GET', 'POST'])
@limiter.exempt
@login_required
@admin_or_staff_required
def profile():
    if request.method == 'POST':
        fname = request.form.get('fname', '').strip()
        lname = request.form.get('lname', '').strip()
        email = request.form.get('email', '').strip()

        if not fname or not lname or not email:
            flash('All fields are required.', 'danger')
            return redirect(url_for('admin.profile'))

        existing_user = User.query.filter_by(email=email).first()
        if existing_user and existing_user.id != current_user.id:
            flash('That email address is already in use.', 'danger')
            return redirect(url_for('admin.profile'))

        current_user.first_name = fname
        current_user.last_name = lname
        current_user.email = email

        try:
            db.session.commit()
            flash('Profile updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash('An error occurred while updating your information.', 'danger')
            
        return redirect(url_for('admin.profile')) 

    return render_template("admin/profile.html")


@admin_bp.route('/profile/update-avatar', methods=['POST'])
@login_required
@admin_or_staff_required
def update_avatar():
    """Handles profile picture uploads via AJAX."""
    if 'avatar' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
        
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400

    if file and allowed_file(file.filename):
        upload_folder = os.path.join(current_app.root_path, 'uploads', 'profiles')
        os.makedirs(upload_folder, exist_ok=True)

        file_ext = file.filename.rsplit('.', 1)[1].lower()
        filename = secure_filename(f"user_{current_user.id}.{file_ext}")
        filepath = os.path.join(upload_folder, filename)

        try:
            # FIXED: Checking and clearing out old image using profile_path
            if current_user.profile_path:
                old_path = os.path.join(upload_folder, current_user.profile_path)
                if os.path.exists(old_path):
                    os.remove(old_path)

            file.save(filepath)
            current_user.profile_path = filename # FIXED: Assigned to database column property
            db.session.commit()

            img_url = url_for('admin.serve_profile_pic', filename=filename)
            return jsonify({'success': True, 'img_url': img_url})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': 'Database error or saving failed.'}), 500

    return jsonify({'success': False, 'message': 'Invalid file extension.'}), 400


@admin_bp.route('/profile/remove-avatar', methods=['POST'])
@login_required
@admin_or_staff_required
def remove_avatar():
    """Handles removing profile picture via AJAX."""
    if not current_user.profile_path:
        return jsonify({'success': True})

    upload_folder = os.path.join(current_app.root_path, 'uploads', 'profiles')
    try:
        old_path = os.path.join(upload_folder, current_user.profile_path)
        if os.path.exists(old_path):
            os.remove(old_path)

        current_user.profile_path = None # FIXED: Cleared database column property
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Could not remove image.'}), 500


@admin_bp.route('/uploads/profiles/<filename>')
@login_required
@limiter.exempt
def serve_profile_pic(filename):
    """Securely serves files from the uploads/profiles folder."""
    return send_from_directory(os.path.join(current_app.root_path, 'uploads', 'profiles'), filename)


@admin_bp.route('/reports')
@login_required
@admin_or_staff_required
@permission_required('can_view_reports')
def reports():
    from datetime import date
    from sqlalchemy import extract
 
    today = date.today()
    current_year = today.year
    current_month = today.month
 
    # ── SALES ──────────────────────────────────────────────────────────────
    total_sales_revenue = db.session.query(
        func.sum(Transaction.amount_paid)
    ).filter(Transaction.transaction_type == 'Sale').scalar() or 0
 
    total_sale_transactions = Transaction.query.filter_by(
        transaction_type='Sale'
    ).count()
 
    cancelled_orders = Transaction.query.filter_by(
        status='Cancelled'
    ).count()
 
    avg_transaction_value = round(
        float(total_sales_revenue) / total_sale_transactions, 2
    ) if total_sale_transactions > 0 else 0
 
    # Sales per month (last 6 months)
    sales_by_month = []
    sales_labels = []
    for i in range(5, -1, -1):
        from dateutil.relativedelta import relativedelta as rd
        target = today - rd(months=i)
        month_total = db.session.query(
            func.sum(Transaction.amount_paid)
        ).filter(
            Transaction.transaction_type == 'Sale',
            extract('year', Transaction.created_at) == target.year,
            extract('month', Transaction.created_at) == target.month
        ).scalar() or 0
        sales_by_month.append(float(month_total))
        sales_labels.append(target.strftime('%b'))
 
    # Top selling products by revenue
    top_products = db.session.query(
        Product.name,
        Product.equipment_type,
        func.sum(Purchase.quantity).label('units_sold'),
        func.sum(Purchase.total_price).label('revenue'),
        Product.stock
    ).join(Purchase, Purchase.product_id == Product.id)\
     .group_by(Product.id)\
     .order_by(func.sum(Purchase.total_price).desc())\
     .limit(5).all()
 
    # Sales by equipment type (for donut chart)
    sales_by_category = db.session.query(
        Product.equipment_type,
        func.sum(Purchase.total_price).label('revenue')
    ).join(Purchase, Purchase.product_id == Product.id)\
     .group_by(Product.equipment_type)\
     .order_by(func.sum(Purchase.total_price).desc())\
     .limit(5).all()
 
    # ── RENTALS ────────────────────────────────────────────────────────────
    active_rentals = Rental.query.filter_by(status='Active').count()
 
    rental_revenue = db.session.query(
        func.sum(Payment.amount)
    ).join(Transaction).filter(
        Transaction.transaction_type == 'Rental',
        Payment.status == 'Completed'
    ).scalar() or 0
 
    overdue_rentals = Rental.query.filter(
        Rental.status == 'Active',
        Rental.expected_return_date < today
    ).all()
 
    returned_rentals = Rental.query.filter_by(status='Returned').count()
 
    # Rentals per month (last 6 months)
    rentals_by_month = []
    for i in range(5, -1, -1):
        target = today - rd(months=i)
        month_count = Rental.query.filter(
            extract('year', Rental.created_at) == target.year,
            extract('month', Rental.created_at) == target.month
        ).count()
        rentals_by_month.append(month_count)
 
    # ── INVENTORY ──────────────────────────────────────────────────────────
    total_inventory = db.session.query(
        func.sum(Product.stock)
    ).scalar() or 0
 
    low_stock_products = Product.query.filter(
        Product.stock <= 5,
        Product.stock > 0
    ).order_by(Product.stock.asc()).all()
 
    out_of_stock_count = Product.query.filter_by(stock=0).count()
 
    stock_health = round(
        (1 - (out_of_stock_count / total_inventory)) * 100, 1
    ) if total_inventory > 0 else 100
 
    # Stock by category
    stock_by_category = db.session.query(
        Product.equipment_type,
        func.sum(Product.stock).label('total_stock')
    ).group_by(Product.equipment_type)\
     .order_by(func.sum(Product.stock).desc())\
     .limit(6).all()
 
    # Most used items (by purchase count)
    most_used = db.session.query(
        Product.name,
        func.count(Purchase.id).label('use_count')
    ).join(Purchase, Purchase.product_id == Product.id)\
     .group_by(Product.id)\
     .order_by(func.count(Purchase.id).desc())\
     .limit(5).all()
 
    max_use = most_used[0].use_count if most_used else 1
 
    # ── CUSTOMERS ──────────────────────────────────────────────────────────
    from models.customer import Customer as CustomerModel
    total_customers = CustomerModel.query.count()
 
    new_customers_this_month = CustomerModel.query.filter(
        extract('year', CustomerModel.created_at) == current_year,
        extract('month', CustomerModel.created_at) == current_month
    ).count()
 
    customers_with_overdue = db.session.query(
        func.count(func.distinct(Rental.customer_id))
    ).filter(
        Rental.status == 'Active',
        Rental.expected_return_date < today
    ).scalar() or 0
 
    # Customer growth per month
    customer_growth = []
    for i in range(5, -1, -1):
        target = today - rd(months=i)
        count = CustomerModel.query.filter(
            extract('year', CustomerModel.created_at) == target.year,
            extract('month', CustomerModel.created_at) == target.month
        ).count()
        customer_growth.append(count)
 
    # Top customers by spending
    top_customers = db.session.query(
        CustomerModel.first_name,
        CustomerModel.last_name,
        func.count(Transaction.id).label('txn_count'),
        func.sum(Transaction.amount_paid).label('total_spent'),
        func.max(Transaction.created_at).label('last_activity')
    ).join(Transaction, Transaction.customer_id == CustomerModel.id)\
     .group_by(CustomerModel.id)\
     .order_by(func.sum(Transaction.amount_paid).desc())\
     .limit(5).all()
 
    # Returning customers (transacted more than once)
    returning_count = db.session.query(
        func.count()
    ).select_from(
        db.session.query(Transaction.customer_id)
        .group_by(Transaction.customer_id)
        .having(func.count(Transaction.id) > 1)
        .subquery()
    ).scalar() or 0
 
    returning_rate = round(
        (returning_count / total_customers) * 100, 1
    ) if total_customers > 0 else 0
 
    # ── FINANCIAL ──────────────────────────────────────────────────────────
    # Gross Revenue = completed sales revenue + completed rental revenue
    sales_revenue_total = db.session.query(
        func.sum(Payment.amount)
    ).join(Transaction).filter(
        Transaction.transaction_type == 'Sale',
        Payment.status == 'Completed'
    ).scalar() or Decimal('0')
 
    rental_revenue_total = db.session.query(
        func.sum(Payment.amount)
    ).join(Transaction).filter(
        Transaction.transaction_type == 'Rental',
        Payment.status == 'Completed'
    ).scalar() or Decimal('0')
 
    gross_revenue = float(sales_revenue_total) + float(rental_revenue_total)
 
    # Cost of Sales = expenses in COST_OF_SALES_CATEGORIES
    cos_rows = db.session.query(
        Expense.category,
        func.sum(Expense.amount).label('total')
    ).filter(
        Expense.category.in_(list(COST_OF_SALES_CATEGORIES))
    ).group_by(Expense.category)\
     .order_by(func.sum(Expense.amount).desc()).all()
 
    cost_of_sales = float(sum(row.total for row in cos_rows) or 0)
 
    # Gross Profit
    gross_profit = gross_revenue - cost_of_sales
 
    # Operating Expenses = expenses in OPERATING_EXPENSE_CATEGORIES
    opex_rows = db.session.query(
        Expense.category,
        func.sum(Expense.amount).label('total')
    ).filter(
        Expense.category.in_(list(OPERATING_EXPENSE_CATEGORIES))
    ).group_by(Expense.category)\
     .order_by(func.sum(Expense.amount).desc()).all()
 
    operating_expenses = float(sum(row.total for row in opex_rows) or 0)
 
    # Net Profit and Profit Margin
    net_profit = gross_profit - operating_expenses
 
    profit_margin = round(
        (net_profit / gross_revenue) * 100, 1
    ) if gross_revenue > 0 else 0
 
    # Operating Expense breakdown dict for the table
    opex_breakdown = {row.category: float(row.total) for row in opex_rows}
    opex_total = operating_expenses if operating_expenses > 0 else 1  
 
    opex_breakdown_list = []
    for row in opex_rows:
        opex_breakdown_list.append({
            'category': row.category,
            'amount': float(row.total),
            'pct': round(float(row.total) / opex_total * 100, 1)
        })
 
    # Cost of Sales breakdown
    cos_breakdown_list = []
    cos_total = cost_of_sales if cost_of_sales > 0 else 1
    for row in cos_rows:
        cos_breakdown_list.append({
            'category': row.category,
            'amount': float(row.total),
            'pct': round(float(row.total) / cos_total * 100, 1)
        })
 
    # ── CHART DATA ─────────────────────────────────────────────────────────
    # Monthly: Revenue, Cost of Sales, Operating Expenses, Net Profit (last 6 months)
    revenue_by_month    = []
    cos_by_month        = []
    opex_by_month       = []
    net_profit_by_month = []
    expenses_by_month   = []   
 
    for i in range(5, -1, -1):
        target = today - rd(months=i)
 
        rev_m = db.session.query(func.sum(Payment.amount)).join(Transaction).filter(
            Payment.status == 'Completed',
            extract('year',  Payment.created_at) == target.year,
            extract('month', Payment.created_at) == target.month,
        ).scalar() or 0
 
        cos_m = db.session.query(func.sum(Expense.amount)).filter(
            Expense.category.in_(list(COST_OF_SALES_CATEGORIES)),
            extract('year',  Expense.date_incurred) == target.year,
            extract('month', Expense.date_incurred) == target.month,
        ).scalar() or 0
 
        opex_m = db.session.query(func.sum(Expense.amount)).filter(
            Expense.category.in_(list(OPERATING_EXPENSE_CATEGORIES)),
            extract('year',  Expense.date_incurred) == target.year,
            extract('month', Expense.date_incurred) == target.month,
        ).scalar() or 0
 
        exp_m = float(cos_m) + float(opex_m)
 
        revenue_by_month.append(float(rev_m))
        cos_by_month.append(float(cos_m))
        opex_by_month.append(float(opex_m))
        net_profit_by_month.append(round(float(rev_m) - exp_m, 2))
        expenses_by_month.append(exp_m)
 
    # Expense breakdown by category (all categories, for donut)
    expense_by_category = db.session.query(
        Expense.category,
        func.sum(Expense.amount).label('total')
    ).group_by(Expense.category)\
     .order_by(func.sum(Expense.amount).desc())\
     .all()
 
    # COS vs OPEX for stacked/grouped chart
    cos_vs_opex_labels  = ['Cost of Sales', 'Operating Expenses']
    cos_vs_opex_data    = [cost_of_sales, operating_expenses]
 
    # Recent transactions for financial tab
    recent_transactions = Transaction.query.order_by(
        Transaction.created_at.desc()
    ).limit(5).all()
    
    # ── Equipment P&L ─────────────────────────────────────────────────────
    # Get all products
    all_equipment_products = Product.query.filter_by(is_active=True).all()

    equipment_pl = []
    for prod in all_equipment_products:

        # Total acquisition cost from linked expenses
        total_spent = float(
            db.session.query(func.sum(Expense.amount))
            .filter(Expense.product_id == prod.id)
            .scalar() or 0
        )

        # Sale income
        sale_income = float(
            db.session.query(func.sum(Purchase.total_price))
            .filter(Purchase.product_id == prod.id)
            .scalar() or 0
        )

        # Rental income — sum of completed payments linked to rentals of this product
        rental_income = float(
            db.session.query(func.sum(Payment.amount))
            .join(Transaction, Payment.transaction_id == Transaction.id)
            .join(Rental, Rental.transaction_id == Transaction.id)
            .filter(
                Rental.product_id == prod.id,
                Payment.status == 'Completed'
            )
            .scalar() or 0
        )

        # Units acquired — sum of quantities from acquisition expenses
        units_acquired = int(
            db.session.query(func.sum(InventoryLog.quantity))
            .filter(
                InventoryLog.product_id == prod.id,
                InventoryLog.action.in_(['Initial Stock Entry', 'Restock'])
            )
            .scalar() or 0
        )

        total_income = sale_income + rental_income
        net = total_income - total_spent

        # Only include products that have some data
        if total_spent > 0 or total_income > 0:
            equipment_pl.append({
                'name': prod.name,
                'equipment_type': prod.equipment_type,
                'units_acquired': units_acquired,
                'total_spent': total_spent,
                'sale_income': sale_income,
                'rental_income': rental_income,
                'total_income': total_income,
                'net': net,
            })

    # Sort by total income descending
    equipment_pl.sort(key=lambda x: x['total_income'], reverse=True)
 
    return render_template(
        "admin/reports.html",
 
        # Sales
        total_sales_revenue=total_sales_revenue,
        total_sale_transactions=total_sale_transactions,
        avg_transaction_value=avg_transaction_value,
        cancelled_orders=cancelled_orders,
        sales_by_month=sales_by_month,
        sales_labels=sales_labels,
        top_products=top_products,
        sales_by_category=sales_by_category,
 
        # Rentals
        active_rentals=active_rentals,
        rental_revenue=rental_revenue,
        overdue_rentals=overdue_rentals,
        returned_rentals=returned_rentals,
        rentals_by_month=rentals_by_month,
 
        # Inventory
        total_inventory=total_inventory,
        low_stock_products=low_stock_products,
        out_of_stock_count=out_of_stock_count,
        stock_health=stock_health,
        stock_by_category=stock_by_category,
        most_used=most_used,
        max_use=max_use,
 
        # Customers
        total_customers=total_customers,
        new_customers_this_month=new_customers_this_month,
        customers_with_overdue=customers_with_overdue,
        returning_rate=returning_rate,
        customer_growth=customer_growth,
        top_customers=top_customers,
        returning_count=returning_count,
 
        # Financial — summary
        gross_revenue=gross_revenue,
        cost_of_sales=cost_of_sales,
        gross_profit=gross_profit,
        operating_expenses=operating_expenses,
        net_profit=net_profit,
        profit_margin=profit_margin,
 
        # Financial — breakdowns
        opex_breakdown_list=opex_breakdown_list,
        cos_breakdown_list=cos_breakdown_list,
 
        # Financial — charts
        revenue_by_month=revenue_by_month,
        cos_by_month=cos_by_month,
        opex_by_month=opex_by_month,
        net_profit_by_month=net_profit_by_month,
        expenses_by_month=expenses_by_month,
        expense_by_category=expense_by_category,
        cos_vs_opex_labels=cos_vs_opex_labels,
        cos_vs_opex_data=cos_vs_opex_data,
 
        # Shared
        recent_transactions=recent_transactions,
        month_labels=sales_labels,
        total_expenses=cost_of_sales + operating_expenses,  # kept for compat
 
        # Chart datasets serialisable
        sales_by_month_json=sales_by_month,
        
        # Equipment P&L
        equipment_pl=equipment_pl,
    )

# Expenses ════════════════════════════════════════════════════════════════════════

@admin_bp.route('/expenses')
@login_required
@admin_or_staff_required
@permission_required('can_manage_expenses')
def expenses():
    from sqlalchemy import extract
    from datetime import date, datetime as dt
 
    today = date.today()
 
    # ── Date range mode ─────────────────────────────────────────────────────
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str   = request.args.get('end_date', '').strip()
    use_range = bool(start_date_str and end_date_str)
 
    start_date = end_date = None
    if use_range:
        try:
            start_date = dt.strptime(start_date_str, '%Y-%m-%d').date()
            end_date   = dt.strptime(end_date_str, '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
        except ValueError:
            use_range = False
 
    # ── Month/Year mode (default) ───────────────────────────────────────────
    year  = request.args.get('year',  today.year,  type=int)
    month = request.args.get('month', today.month, type=int)
    category_filter = request.args.get('category', '')
 
    # ── Build expense query ─────────────────────────────────────────────────
    query = Expense.query
 
    if use_range:
        query = query.filter(
            Expense.date_incurred >= start_date,
            Expense.date_incurred <= end_date,
        )
    else:
        query = query.filter(
            extract('year',  Expense.date_incurred) == year,
            extract('month', Expense.date_incurred) == month,
        )
 
    if category_filter:
        query = query.filter(Expense.category == category_filter)
 
    expenses_list = query.order_by(Expense.date_incurred.desc()).all()
 
    # ── Category totals (for charts) ────────────────────────────────────────
    def _period_filter(q):
        if use_range:
            return q.filter(
                Expense.date_incurred >= start_date,
                Expense.date_incurred <= end_date,
            )
        return q.filter(
            extract('year',  Expense.date_incurred) == year,
            extract('month', Expense.date_incurred) == month,
        )
 
    cat_query = _period_filter(
        db.session.query(
            Expense.category,
            func.sum(Expense.amount).label('total')
        )
    ).group_by(Expense.category)\
     .order_by(func.sum(Expense.amount).desc())
 
    category_totals = cat_query.all()
 
    # ── Financial metrics for the period ────────────────────────────────────
    # Cost of Sales
    cos_total = float(
        _period_filter(
            db.session.query(func.sum(Expense.amount))
            .filter(Expense.category.in_(list(COST_OF_SALES_CATEGORIES)))
        ).scalar() or 0
    )
 
    # Operating Expenses
    opex_total = float(
        _period_filter(
            db.session.query(func.sum(Expense.amount))
            .filter(Expense.category.in_(list(OPERATING_EXPENSE_CATEGORIES)))
        ).scalar() or 0
    )
 
    total_expenses = cos_total + opex_total
 
    # Gross Revenue for the period (completed payments)
    rev_query = db.session.query(func.sum(Payment.amount)).join(Transaction).filter(
        Payment.status == 'Completed'
    )
    if use_range:
        rev_query = rev_query.filter(
            Payment.created_at >= dt.combine(start_date, dt.min.time()),
            Payment.created_at <= dt.combine(end_date, dt.max.time()),
        )
    else:
        rev_query = rev_query.filter(
            extract('year',  Payment.created_at) == year,
            extract('month', Payment.created_at) == month,
        )
    gross_revenue = float(rev_query.scalar() or 0)
 
    gross_profit     = gross_revenue - cos_total
    net_profit       = gross_profit  - opex_total
    profit_margin    = round((net_profit / gross_revenue) * 100, 1) if gross_revenue > 0 else 0
 
    # ── Last 6-month trend (always month-based) ─────────────────────────────
    from dateutil.relativedelta import relativedelta as rd
    trend_labels   = []
    trend_expense  = []   # total expenses (COS + OPEX)
    trend_cos      = []
    trend_opex     = []
    trend_revenue  = []
    trend_net      = []
 
    for i in range(5, -1, -1):
        target = today - rd(months=i)
        exp = float(db.session.query(func.sum(Expense.amount)).filter(
            extract('year',  Expense.date_incurred) == target.year,
            extract('month', Expense.date_incurred) == target.month,
        ).scalar() or 0)
        cos_m = float(db.session.query(func.sum(Expense.amount)).filter(
            Expense.category.in_(list(COST_OF_SALES_CATEGORIES)),
            extract('year',  Expense.date_incurred) == target.year,
            extract('month', Expense.date_incurred) == target.month,
        ).scalar() or 0)
        opex_m = float(db.session.query(func.sum(Expense.amount)).filter(
            Expense.category.in_(list(OPERATING_EXPENSE_CATEGORIES)),
            extract('year',  Expense.date_incurred) == target.year,
            extract('month', Expense.date_incurred) == target.month,
        ).scalar() or 0)
        rev = float(db.session.query(func.sum(Payment.amount)).join(Transaction).filter(
            Payment.status == 'Completed',
            extract('year',  Payment.created_at) == target.year,
            extract('month', Payment.created_at) == target.month,
        ).scalar() or 0)
        trend_labels.append(target.strftime('%b %Y'))
        trend_expense.append(exp)
        trend_cos.append(cos_m)
        trend_opex.append(opex_m)
        trend_revenue.append(rev)
        trend_net.append(round(rev - exp, 2))
 
    # ── Category type lookup (for table badges) ─────────────────────────────
    def category_type(cat):
        if cat in COST_OF_SALES_CATEGORIES:
            return 'cos'
        if cat in OPERATING_EXPENSE_CATEGORIES:
            return 'opex'
        return 'other'
 
    months = [
        (1,'January'),(2,'February'),(3,'March'),(4,'April'),
        (5,'May'),(6,'June'),(7,'July'),(8,'August'),
        (9,'September'),(10,'October'),(11,'November'),(12,'December')
    ]
 
    return render_template(
        'admin/expenses.html',
        # Expense records
        expenses=expenses_list,
        category_totals=category_totals,
        category_type=category_type,
 
        # Period totals
        total_this_month=total_expenses,
        revenue_this_month=gross_revenue,
 
        # Financial summary
        gross_revenue=gross_revenue,
        cost_of_sales=cos_total,
        gross_profit=gross_profit,
        operating_expenses=opex_total,
        net_profit=net_profit,
        profit_margin=profit_margin,
 
        # Trend charts
        trend_labels=trend_labels,
        trend_expense=trend_expense,
        trend_cos=trend_cos,
        trend_opex=trend_opex,
        trend_revenue=trend_revenue,
        trend_net=trend_net,
 
        # Filter state
        all_categories=ALL_EXPENSE_CATEGORIES_ORDERED,
        category_filter=category_filter,
        months=months,
        current_year=year,
        current_month=month,
        use_range=use_range,
        start_date=start_date,
        end_date=end_date,
        today=today,
 
        # Sets for Jinja checks
        cos_categories=COST_OF_SALES_CATEGORIES,
        opex_categories=OPERATING_EXPENSE_CATEGORIES,
    )

@admin_bp.route('/expenses/add', methods=['POST'])
@login_required
@admin_or_staff_required
def add_expense():
    if current_user.role.strip() != 'Administrator':
        flash("Only Administrators can record expenses.", "danger")
        return redirect(url_for('admin.expenses'))
    
    from datetime import datetime
    try:
        category    = request.form.get('category', '').strip()
        expense_title = request.form.get('expense_title', '').strip()
        amount_raw  = request.form.get('amount', '0').replace(',', '').strip()
        description = request.form.get('description', '').strip()
        date_str    = request.form.get('date_incurred', '').strip()
 
        if not category or not amount_raw or not expense_title:
            flash('Category, expense name, and amount are required.', 'error')
            return redirect(request.referrer)
 
        amount = Decimal(amount_raw)
        if amount <= 0:
            flash('Amount must be greater than zero.', 'error')
            return redirect(request.referrer)
 
        date_incurred = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()
 
        # Handle receipt attachment
        attachment_path = None
        file = request.files.get('attachment')
        if file and file.filename:
            ext = os.path.splitext(secure_filename(file.filename))[1].lower()
            if ext not in ['.jpg', '.jpeg', '.png', '.webp', '.pdf']:
                flash('Invalid file type. Use JPG, PNG, WebP, or PDF.', 'error')
                return redirect(request.referrer)
            filename = f"exp_{uuid.uuid4().hex[:12]}{ext}"
            folder = os.path.join(current_app.root_path, 'static', 'uploads', 'expenses')
            os.makedirs(folder, exist_ok=True)
            file.save(os.path.join(folder, filename))
            attachment_path = f'uploads/expenses/{filename}'
 
        new_expense = Expense(
            category=category,
            expense_title=expense_title,
            amount=amount,
            description=description,
            date_incurred=date_incurred,
            attachment_path=attachment_path,
            recorded_by_id=current_user.id,
        )
        db.session.add(new_expense)
 
        log = InventoryLog(
            action='Expense Recorded',
            note=f'{expense_title} ({category}) — ₱{amount:,.2f}: {description[:80] if description else "No notes"}',
            user_id=current_user.id,
            user_name=current_user.full_name,
        )
        db.session.add(log)
        db.session.commit()
 
        flash(f'Expense of ₱{amount:,.2f} recorded successfully.', 'success')
 
    except (InvalidOperation, ValueError):
        db.session.rollback()
        flash('Invalid amount format.', 'error')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'ADD_EXPENSE_ERROR: {e}')
        flash('An error occurred while saving the expense.', 'error')
 
    return redirect(request.referrer or url_for('admin.expenses',
                                             year=request.form.get('year', ''),
                                             month=request.form.get('month', '')))
 
 
@admin_bp.route('/expenses/delete/<int:expense_id>', methods=['POST'])
@login_required
@admin_or_staff_required
def delete_expense(expense_id):
    if current_user.role.strip() != 'Administrator':
        flash("Only Administrators can delete expenses.", "danger")
        return redirect(url_for('admin.expenses'))
    
    expense = Expense.query.get_or_404(expense_id)
    try:
        # Delete attachment file if it exists
        if expense.attachment_path:
            full_path = os.path.join(current_app.root_path, 'static', expense.attachment_path)
            if os.path.exists(full_path):
                os.remove(full_path)
 
        log = InventoryLog(
            action='Expense Deleted',
            note=f'Deleted expense: {expense.category} — ₱{expense.amount:,.2f}',
            user_id=current_user.id,
            user_name=current_user.full_name,
        )
        db.session.add(log)
        db.session.delete(expense)
        db.session.commit()
        flash('Expense deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'DELETE_EXPENSE_ERROR: {e}')
        flash('Failed to delete expense.', 'error')
 
    return redirect(request.referrer or url_for('admin.expenses'))
 
 
@admin_bp.route('/expenses/edit/<int:expense_id>', methods=['POST'])
@login_required
@admin_or_staff_required
def edit_expense(expense_id):
    if current_user.role.strip() != 'Administrator':
        flash("Only Administrators can edit expenses.", "danger")
        return redirect(url_for('admin.expenses'))
    
    from datetime import datetime
    expense = Expense.query.get_or_404(expense_id)
    try:
        expense.category      = request.form.get('category', expense.category).strip()
        expense.expense_title = request.form.get('expense_title', expense.expense_title).strip()
        expense.description   = request.form.get('description', '').strip()
        date_str = request.form.get('date_incurred', '')
        if date_str:
            expense.date_incurred = datetime.strptime(date_str, '%Y-%m-%d').date()
 
        amount_raw = request.form.get('amount', '0').replace(',', '').strip()
        expense.amount = Decimal(amount_raw)
 
        db.session.commit()
        flash('Expense updated.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'EDIT_EXPENSE_ERROR: {e}')
        flash('Failed to update expense.', 'error')
 
    return redirect(request.referrer or url_for('admin.expenses'))


# ── Security Dashboard 
@admin_bp.route('/security')
@limiter.exempt
@login_required
def security_dashboard():
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))

    logs = SecurityLog.query.order_by(SecurityLog.created_at.desc()).limit(500).all()
    blocked = BlockedIP.query.filter_by(is_active=True).order_by(BlockedIP.blocked_at.desc()).all()
 
    total_logs        = SecurityLog.query.count()
    suspicious_count  = SecurityLog.query.filter_by(is_suspicious=True).count()
    blocked_count     = BlockedIP.query.filter_by(is_active=True).count()
    rate_limit_count  = SecurityLog.query.filter_by(event_type="Rate Limit Violation").count()
 
    failed_count      = SecurityLog.query.filter(SecurityLog.event_type.ilike("%failed%")).count()
    locked_count      = SecurityLog.query.filter(SecurityLog.event_type.ilike("%lock%")).count()
    blocked_ip_count  = SecurityLog.query.filter(SecurityLog.event_type.ilike("%block%")).count()
    bot_count         = SecurityLog.query.filter_by(event_type="Bot Detected").count()
 
    # ── Admin Activity tab ────────────────────────────────────────────────
    # All event types that belong to admin activity monitoring
    admin_event_types = [
        'Admin Login',
        'Admin Login — New Device',
        'Admin Logout',
        'Admin Session Expired',
        'Admin Failed Login',
    ]

    staff_event_types = [
        'Staff Login',
        'Staff Login — New Device',
        'Staff Logout',
        'Staff Session Expired',
        'Staff Failed Login',
        'Staff Action',
    ]
 
    admin_logs = SecurityLog.query.filter(
        SecurityLog.event_type.in_(admin_event_types)
    ).order_by(SecurityLog.created_at.desc()).limit(200).all()

    admin_activity_count = SecurityLog.query.filter(
        SecurityLog.event_type.in_(admin_event_types)
    ).count()

    admin_login_count = SecurityLog.query.filter(
        SecurityLog.event_type.in_(['Admin Login', 'Admin Login — New Device'])
    ).count()

    admin_logout_count = SecurityLog.query.filter_by(
        event_type='Admin Logout'
    ).count()

    admin_expired_count = SecurityLog.query.filter_by(
        event_type='Admin Session Expired'
    ).count()

    admin_failed_count = SecurityLog.query.filter_by(
        event_type='Admin Failed Login'
    ).count()

    # ── Staff Activity ─────────────────────────────────────────────────────
    staff_logs = SecurityLog.query.filter(
        SecurityLog.event_type.in_(staff_event_types)
    ).order_by(SecurityLog.created_at.desc()).limit(200).all()

    staff_activity_count = SecurityLog.query.filter(
        SecurityLog.event_type.in_(staff_event_types)
    ).count()

    staff_login_count = SecurityLog.query.filter(
        SecurityLog.event_type.in_(['Staff Login', 'Staff Login — New Device'])
    ).count()

    staff_logout_count = SecurityLog.query.filter_by(
        event_type='Staff Logout'
    ).count()

    staff_action_count = SecurityLog.query.filter_by(
        event_type='Staff Action'
    ).count()

    # ── All staff actions from InventoryLog (what they did) ───────────────
    staff_users = User.query.filter_by(role='Staff').all()
    staff_user_ids = [u.id for u in staff_users]

    staff_activity_logs = InventoryLog.query.filter(
        InventoryLog.user_id.in_(staff_user_ids)
    ).order_by(InventoryLog.created_at.desc()).limit(200).all()
 
    return render_template(
        'admin/security_dashboard.html',
        logs=logs,
        blocked=blocked,
        total_logs=total_logs,
        suspicious_count=suspicious_count,
        blocked_count=blocked_count,
        rate_limit_count=rate_limit_count,
        failed_count=failed_count,
        locked_count=locked_count,
        blocked_ip_count=blocked_ip_count,
        bot_count=bot_count,
        # admin activity
        admin_logs=admin_logs,
        admin_activity_count=admin_activity_count,
        admin_login_count=admin_login_count,
        admin_logout_count=admin_logout_count,
        admin_expired_count=admin_expired_count,
        admin_failed_count=admin_failed_count,
        # staff activity
        staff_logs=staff_logs,
        staff_activity_count=staff_activity_count,
        staff_login_count=staff_login_count,
        staff_logout_count=staff_logout_count,
        staff_action_count=staff_action_count,
        staff_activity_logs=staff_activity_logs,
        staff_members=staff_users,
    )


# ── Block an IP (manual or auto via JS) 
@admin_bp.route('/security/block-ip', methods=['POST'])
@login_required
@admin_or_staff_required
def block_ip():
    ip = request.form.get('ip_address', '').strip()
    reason = request.form.get('reason', 'Manually blocked by admin').strip()
    duration_hours = request.form.get('duration_hours', '').strip()
 
    if not ip:
        flash("IP address is required.", "error")
        return redirect(url_for('admin.security_dashboard'))
 
    from datetime import datetime, timedelta
 
    blocked_until = None
    if duration_hours:
        try:
            h = int(duration_hours)
            if h > 0:
                blocked_until = datetime.utcnow() + timedelta(hours=h)
        except ValueError:
            pass
 
    existing = BlockedIP.query.filter_by(ip_address=ip).first()
    if existing:
        existing.is_active = True
        existing.reason = reason
        existing.blocked_until = blocked_until
        existing.blocked_at = datetime.utcnow()
        existing.blocked_by = current_user.id
    else:
        entry = BlockedIP(
            ip_address=ip,
            reason=reason,
            blocked_until=blocked_until,
            is_active=True,
            blocked_by=current_user.id
        )
        db.session.add(entry)
 
    # Log the manual block in security logs
    log = SecurityLog(
        ip_address=ip,
        event_type="IP Blocked",
        description=f"Admin manually blocked IP. Reason: {reason}",
        user_id=current_user.id,
        user_email=current_user.email,
        is_suspicious=True
    )
    db.session.add(log)
 
    try:
        db.session.commit()
        flash(f"IP {ip} has been blocked successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error blocking IP: {str(e)}", "error")
 
    return redirect(url_for('admin.security_dashboard'))
 
 
# ── Unblock an IP 
@admin_bp.route('/security/unblock/<int:block_id>', methods=['POST'])
@login_required
@admin_or_staff_required
def unblock_ip(block_id):
    entry = BlockedIP.query.get_or_404(block_id)
    entry.is_active = False
 
    log = SecurityLog(
        ip_address=entry.ip_address,
        event_type="IP Unblocked",
        description=f"Admin unblocked IP {entry.ip_address}",
        user_id=current_user.id,
        user_email=current_user.email,
        is_suspicious=False
    )
    db.session.add(log)
 
    try:
        db.session.commit()
        flash(f"IP {entry.ip_address} has been unblocked.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error unblocking IP: {str(e)}", "error")
 
    return redirect(url_for('admin.security_dashboard'))
 
 
# ── Delete a single security log 
@admin_bp.route('/security/delete-log/<int:log_id>', methods=['POST'])
@login_required
@admin_or_staff_required
def delete_security_log(log_id):
    log = SecurityLog.query.get_or_404(log_id)
    try:
        db.session.delete(log)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
 
 
# ── Clear ALL security logs 
@admin_bp.route('/security/clear-logs', methods=['POST'])
@login_required
@admin_or_staff_required
def clear_security_logs():
    try:
        SecurityLog.query.delete()
        db.session.commit()
        flash("All security logs have been cleared.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error clearing logs: {str(e)}", "error")
    return redirect(url_for('admin.security_dashboard'))

# ── Backup Management ──────────────────────────
@admin_bp.route('/backup')
@login_required
def backup_page():
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    backups = get_all_backups()
    total_size = sum(b['size_kb'] for b in backups)
    return render_template('admin/backup.html', backups=backups, total_size=round(total_size, 2))


@admin_bp.route('/backup/create', methods=['POST'])
@login_required
def create_manual_backup():
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    success, result = create_backup(triggered_by='admin')

    try:
        log = SecurityLog(
            ip_address=request.remote_addr,
            event_type="Manual Backup Created" if success else "Backup Failed",
            description=f"Admin triggered manual backup. Result: {result}",
            user_id=current_user.id,
            user_email=current_user.email,
            user_agent=request.headers.get('User-Agent', 'Unknown')[:255],
            severity='Low',
            is_suspicious=False
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()

    if success:
        flash(f"Backup created successfully: {result}", "success")
    else:
        flash(f"Backup failed: {result}", "error")

    return redirect(url_for('admin.backup_page'))


@admin_bp.route('/backup/download/<filename>')
@login_required
def download_backup(filename):
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    # Prevent path traversal attack
    if '..' in filename or '/' in filename or '\\' in filename:
        flash("Invalid filename.", "error")
        return redirect(url_for('admin.backup_page'))

    filepath = os.path.join('backups', filename)
    if not os.path.exists(filepath):
        flash("Backup file not found.", "error")
        return redirect(url_for('admin.backup_page'))

    try:
        log = SecurityLog(
            ip_address=request.remote_addr,
            event_type="Backup Downloaded",
            description=f"Admin downloaded backup: {filename}",
            user_id=current_user.id,
            user_email=current_user.email,
            severity='Low',
            is_suspicious=False
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()

    return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/sql')


@admin_bp.route('/backup/delete/<filename>', methods=['POST'])
@login_required
def delete_backup(filename):
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    if '..' in filename or '/' in filename or '\\' in filename:
        flash("Invalid filename.", "error")
        return redirect(url_for('admin.backup_page'))

    filepath = os.path.join('backups', filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        flash(f"Backup deleted successfully.", "success")
    else:
        flash("Backup file not found.", "error")

    return redirect(url_for('admin.backup_page'))


def _build_notifications():

    from models.product import PaymentProof, RentalInvoice
    from models.customer import Customer
    from datetime import timedelta, date

    today = date.today()
    soon  = today + timedelta(days=5)
    notifications = []

    # ── 1. RENTALS WITH OVERDUE PAYMENT ───────────────────────────────────
    overdue_payment_txns = db.session.query(Transaction).join(
        Rental, Rental.transaction_id == Transaction.id
    ).filter(
        Transaction.transaction_type == 'Rental',
        Transaction.balance_due > 0,
        Transaction.status == 'Open',
        Rental.start_date <= today
    ).distinct().all()

    overdue_payment_count = len(overdue_payment_txns)

    if overdue_payment_count > 0:
        notifications.append({
            "id": "overdue_payments",
            "type": "error",
            "icon": "assignment_late",
            "title": f"{overdue_payment_count} Rental{'s' if overdue_payment_count > 1 else ''} with Overdue Payment{'s' if overdue_payment_count > 1 else ''}",
            "message": "Click to view and collect outstanding rental payments.",
            "link": "/admin/transactions?type=Rental&status=overdue_payment"
        })

    # ── 2. RENTALS EXPIRING WITHIN 3 DAYS ─────────────────────────────────
    expiring_soon = Rental.query.filter(
        Rental.status == 'Active',
        Rental.expected_return_date >= today,
        Rental.expected_return_date <= soon
    ).all()
    expiring_count = len(expiring_soon)

    if expiring_count > 0:
        notifications.append({
            "id": "expiring_soon",
            "type": "warning",
            "icon": "event_upcoming",
            "title": f"{expiring_count} Rental{'s' if expiring_count > 1 else ''} Expiring Soon",
            "message": "Click to view rentals expiring within 5 days.",
            "link": "/admin/transactions?type=Rental&status=expiring"
        })

    # ── 3. OVERDUE RETURNS (items not returned past due date) ──────────────────
    overdue_returns = Rental.query.filter(
        Rental.status == 'Active',
        Rental.expected_return_date < today
    ).all()

    if overdue_returns:
        first_item = overdue_returns[0]
        days_overdue = (today - first_item.expected_return_date).days
        
        notifications.append({
            "id": "overdue_returns",
            "type": "error",
            "icon": "assignment_return",
            "title": f"{len(overdue_returns)} Equipment Return{'s' if len(overdue_returns) > 1 else ''} OVERDUE",
            "message": f"{len(overdue_returns)} item{'s' if len(overdue_returns) > 1 else ''} not returned. Oldest: {days_overdue} days late.",
            "link": "/admin/active-rentals?filter=overdue_return" 
        })

    # ── 4. UNPAID SALE TRANSACTIONS ───────────────────────────────────────
    unpaid_sales = Transaction.query.filter(
        Transaction.transaction_type == 'Sale',
        Transaction.payment_status == 'Unpaid',
        Transaction.status == 'Open'
    ).count()

    if unpaid_sales > 0:
        notifications.append({
            "id": "unpaid_sales",
            "type": "warning",
            "icon": "money_off",
            "title": f"{unpaid_sales} Unpaid Sale Transaction{'s' if unpaid_sales > 1 else ''}",
            "message": f"{unpaid_sales} sale{'s' if unpaid_sales > 1 else ''} with outstanding balance.",
            "link": "/admin/transactions?type=Sale&status=unpaid"
        })

    # ── 5. PARTIALLY PAID TRANSACTIONS ────────────────────────────────────
    partial_payments = Transaction.query.filter(
        Transaction.payment_status == 'Partially Paid',
        Transaction.status == 'Open'
    ).count()

    if partial_payments > 0:
        notifications.append({
            "id": "partial_payments",
            "type": "warning",
            "icon": "payments",
            "title": f"{partial_payments} Partially Paid Transaction{'s' if partial_payments > 1 else ''}",
            "message": f"{partial_payments} transaction{'s' if partial_payments > 1 else ''} still have remaining balances.",
            "link": "/admin/transactions?status=partial"
        })

    # ── 6. LOW STOCK ──────────────────────────────────────────────────────
    low_stock = Product.query.filter(
        Product.stock <= 5,
        Product.stock > 0,
        Product.status != 'Archived'
    ).all()

    if low_stock:
        names = ", ".join([p.name for p in low_stock[:3]])
        extra = f" and {len(low_stock) - 3} more" if len(low_stock) > 3 else ""
        notifications.append({
            "id": "low_stock",
            "type": "warning",
            "icon": "inventory_2",
            "title": f"{len(low_stock)} Item{'s' if len(low_stock) > 1 else ''} Running Low",
            "message": f"{names}{extra} — restock soon.",
            "link": "/admin/products?filter=low_stock"
        })

    # ── 7. OUT OF STOCK ───────────────────────────────────────────────────
    out_of_stock = Product.query.filter(
        Product.stock == 0,
        Product.status != 'Archived'
    ).count()

    if out_of_stock > 0:
        notifications.append({
            "id": "out_of_stock",
            "type": "error",
            "icon": "remove_shopping_cart",
            "title": f"{out_of_stock} Item{'s' if out_of_stock > 1 else ''} Out of Stock",
            "message": "Immediate restock required.",
            "link": "/admin/products?filter=out_of_stock"
        })

    # ── 8. ORDERS STUCK AT SUBMITTED ─────────────────────────────────────
    stuck_submitted = Transaction.query.filter(
        Transaction.tracking_status == 'SUBMITTED',
        Transaction.status == 'Open'
    ).count()

    if stuck_submitted > 0:
        notifications.append({
            "id": "stuck_submitted",
            "type": "info",
            "icon": "hourglass_top",
            "title": f"{stuck_submitted} Order{'s' if stuck_submitted > 1 else ''} Awaiting Verification",
            "message": f"{stuck_submitted} order{'s' if stuck_submitted > 1 else ''} submitted and not yet verified.",
            "link": "/admin/transactions?status=submitted"
        })

    # ── 9. PENDING CUSTOMER ID VERIFICATIONS ─────────────────────────────
    pending_verifications = Customer.query.filter(
        Customer.is_id_verified == False,
        Customer.valid_id_path != None
    ).count()

    if pending_verifications > 0:
        notifications.append({
            "id": "pending_verifications",
            "type": "info",
            "icon": "verified_user",
            "title": f"{pending_verifications} ID Verification{'s' if pending_verifications > 1 else ''} Pending",
            "message": "Customer IDs uploaded and awaiting admin review.",
            "link": "/admin/customers?filter=pending_id"
        })

    # ── 10. INACTIVE CUSTOMERS WITH OPEN TRANSACTIONS ─────────────────────
    inactive_with_open = db.session.query(func.count(Transaction.id)).join(
        Customer, Customer.id == Transaction.customer_id
    ).filter(
        Customer.is_active == False,
        Transaction.status == 'Open'
    ).scalar() or 0

    if inactive_with_open > 0:
        notifications.append({
            "id": "inactive_open_txn",
            "type": "warning",
            "icon": "person_off",
            "title": f"{inactive_with_open} Open Transaction{'s' if inactive_with_open > 1 else ''} — Inactive Customer{'s' if inactive_with_open > 1 else ''}",
            "message": "Deactivated customers still have open transactions.",
            "link": "/admin/customers?filter=inactive_open"
        })

    return notifications

@admin_bp.route('/notifications/stream')
@login_required
@admin_or_staff_required
def notification_stream():
    """Server-Sent Events endpoint — pushes notification data every 30s."""
 
    def generate():
        notifications = _build_notifications()
        payload = json.dumps({"count": len(notifications), "notifications": notifications})
        yield f"data: {payload}\n\n"
 
        import time
        interval = 30
        elapsed = 0
        while True:
            time.sleep(1)
            elapsed += 1
            if elapsed >= interval:
                elapsed = 0
                notifications = _build_notifications()
                payload = json.dumps({"count": len(notifications), "notifications": notifications})
                yield f"data: {payload}\n\n"
 
    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'
        }
    )
 
 
@admin_bp.route('/notifications/data')
@login_required
@admin_or_staff_required
def notification_data():
    """One-shot JSON endpoint — used for the initial page load."""
    try:
        notifications = _build_notifications()
        return jsonify({"count": len(notifications), "notifications": notifications})
    except Exception as e:
        current_app.logger.error(f"Notification data error: {e}")
        return jsonify({"count": 0, "notifications": []})
 

@admin_bp.route('/staff-management', methods=['GET', 'POST'])
@login_required
def staff_management():
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    staff_members = User.query.filter(User.role.in_(['Staff', 'Administrator']))\
                              .options(joinedload(User.permissions))\
                              .all()

    total_staff_count = len(staff_members)
    admin_count = sum(1 for u in staff_members if u.role == 'Administrator')
    active_staff_count = sum(1 for u in staff_members if u.is_active)

    return render_template(
        'admin/staff_management.html',
        staff_members=staff_members,
        total_staff_count=total_staff_count,
        admin_count=admin_count,
        active_staff_count=active_staff_count
    )


from extensions import db, passhasher

@admin_bp.route("/staff/create", methods=["POST"])
@login_required
def create_system_user():
    if current_user.role.strip() != 'Administrator':
        return jsonify({"success": False, "message": "Unauthorized."}), 403
    
    data = {
        "email": request.form.get("email", "").strip().lower(),
        "first": request.form.get("first_name", "").strip().title(),
        "last": request.form.get("last_name", "").strip().title(),
        "pwd": request.form.get("password", ""),
        "role": request.form.get("role")
    }

    if not all(data.values()):
        flash("All fields are mandatory.", "danger")
        return redirect(url_for("admin.staff_management"))

    pwd_errors = []
    if len(data['pwd']) < 12: pwd_errors.append("12+ characters")
    if not any(c.isupper() for c in data['pwd']): pwd_errors.append("uppercase letter")
    if not any(c.isdigit() for c in data['pwd']): pwd_errors.append("number")
    if not any(c in "!@#$%^&*()-_=+" for c in data['pwd']): pwd_errors.append("special character")

    if pwd_errors:
        return jsonify({"success": False, "message": f"Password weak. Requires: {', '.join(pwd_errors)}"}), 400

    if User.query.filter_by(email=data['email']).first():
        return jsonify({"success": False, "message": "Email already exists in the system."}), 400

    try:
        user_record = User(
            email=data['email'],
            password_hash=passhasher.hash(data['pwd']),
            first_name=data['first'],
            last_name=data['last'],
            role=data['role'],
            is_active=True,
            is_verified=True
        )
        
        db.session.add(user_record)
        db.session.flush() 


        new_permissions = Permission(user_id=user_record.id)
        db.session.add(new_permissions)

        db.session.commit()

        return jsonify({"success": True, "message": f"New system user {data['first']} initialized successfully."})
        
    except Exception as error:
        db.session.rollback()
        current_app.logger.critical(f"Provisioning Failure: {error}")
        return jsonify({"success": False, "message": "System error during record creation. Contact development team."}), 500
    
@admin_bp.route("/update-user-access", methods=["POST"])
@login_required
def update_user_access():
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    user_id = request.form.get("user_id")
    

    user = User.query.get_or_404(user_id)

    user.role = request.form.get("role")

    perms = Permission.query.filter_by(user_id=user_id).first()
    if not perms:
        perms = Permission(user_id=user_id)
        db.session.add(perms)
    

    perms.can_manage_customers = 'can_manage_customers' in request.form
    perms.can_manage_products = 'can_manage_products' in request.form
    perms.can_process_transactions = 'can_process_transactions' in request.form
    perms.can_confirm_payments = 'can_confirm_payments' in request.form
    perms.can_manage_expenses = 'can_manage_expenses' in request.form
    perms.can_view_reports = 'can_view_reports' in request.form
    perms.can_view_active_rentals = 'can_view_active_rentals' in request.form
    
    try:
        db.session.commit()
        flash(f"Access rights and role updated for {user.full_name}.", "success")
    except Exception as e:
        db.session.rollback()
        flash("An error occurred while updating permissions.", "error")
        
    return redirect(url_for('admin.staff_management'))

@admin_bp.route("/delete-user/<int:user_id>", methods=["POST"])
@login_required
def delete_user(user_id):
    if current_user.role.strip() != 'Administrator':
        flash("Access restricted to Administrators only.", "danger")
        return redirect(url_for('admin.dashboard'))
    
    user = User.query.get_or_404(user_id)

    # Prevent self-deletion
    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for('admin.staff_management'))

    try:
        # Delete associated permissions first to avoid FK constraint errors
        Permission.query.filter_by(user_id=user.id).delete()
        db.session.delete(user)
        db.session.commit()
        flash(f"{user.first_name} {user.last_name} has been removed from the system.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"DELETE_USER_ERROR: {e}")
        flash("Failed to delete user. They may have linked records.", "danger")

    return redirect(url_for('admin.staff_management'))